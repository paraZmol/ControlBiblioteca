# endpoints.py - Rutas de la API REST
import asyncio
from datetime import datetime, date, timedelta
from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, Body, Request
from fastapi.security import OAuth2PasswordRequestForm
from sqlalchemy import select, update, delete
from sqlalchemy.ext.asyncio import AsyncSession
from database import get_db
from models import AlumnoMaestro, Terminal, Sesion, Usuario, Facultad, Escuela, Ban, PersonalUniversidad, Incidencia, Sospecha, ActividadLog, ProcesoIgnorado, Auditoria
Alumno = AlumnoMaestro  # alias local para compatibilidad con código legacy
from auth_service import (
    verificar_password, hashear_password, crear_token, obtener_usuario_actual
)
from core.websocket_manager import manager
from core.auditoria import registrar_auditoria
from core import rate_limit
from pydantic import BaseModel

router = APIRouter(prefix="/api")

# E-10: límite de tamaño para subidas de Excel (DoS por archivo gigante).
# 25 MB cubre de sobra un padrón de alumnos/personal. Configurable por env.
import os as _os_mod
try:
    _MAX_EXCEL_BYTES = int(_os_mod.getenv("MAX_EXCEL_MB", "25")) * 1024 * 1024
except ValueError:
    _MAX_EXCEL_BYTES = 25 * 1024 * 1024


async def _leer_upload_limitado(archivo, max_bytes: int = _MAX_EXCEL_BYTES) -> bytes:
    """Lee un UploadFile en trozos, abortando con HTTP 413 si supera max_bytes (E-10).

    Evita cargar en memoria un archivo arbitrariamente grande: lee por chunks y
    corta apenas se pasa del límite, sin haber materializado el archivo entero.
    """
    buf = bytearray()
    while True:
        chunk = await archivo.read(1024 * 256)  # 256 KB
        if not chunk:
            break
        buf.extend(chunk)
        if len(buf) > max_bytes:
            raise HTTPException(
                status_code=413,
                detail=f"El archivo supera el tamaño máximo permitido ({max_bytes // (1024*1024)} MB)."
            )
    return bytes(buf)


# E-7: hash bcrypt "señuelo" para igualar el tiempo de respuesta cuando el
# usuario NO existe. Sin esto, "usuario inexistente" responde al instante y
# "contraseña incorrecta" tarda ~300ms (bcrypt), lo que permite enumerar
# usuarios válidos por timing. Verificamos siempre contra un hash real para
# gastar el mismo tiempo de CPU exista o no el usuario.
_HASH_SENUELO = hashear_password("timing-attack-mitigation-dummy")


# ── Esquemas Pydantic ──────────────────────────────────────────────

class LoginResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"

class AlumnoAuth(BaseModel):
    codigo: str

class TerminalRegistro(BaseModel):
    nombre: str
    ip: str

class SesionResponse(BaseModel):
    id: int
    alumno_codigo: str
    alumno_nombre: str
    terminal_nombre: str
    inicio: datetime
    activa: bool

class UsuarioCrear(BaseModel):
    username: str
    password: str
    nombre_completo: str | None = None
    rol: str = "admin"


# ── Server Info ────────────────────────────────────────────────────

# ── Auth ────────────────────────────────────────────────────────────

@router.post("/auth/login", response_model=LoginResponse)
async def login(request: Request, form: OAuth2PasswordRequestForm = Depends(), db: AsyncSession = Depends(get_db)):
    """Iniciar sesión como administrador/encargado."""
    from datetime import timedelta
    # E-7: límite por IP además del bloqueo por usuario. Frena la fuerza bruta
    # distribuida (probar muchos usuarios desde una IP). 10 intentos / 5 min.
    rate_limit.exigir(request, "login", limite=10, ventana_seg=300,
                      mensaje="Demasiados intentos de inicio de sesión. Espere unos minutos.")

    result = await db.execute(select(Usuario).where(Usuario.username == form.username))
    usuario = result.scalar_one_or_none()

    # E-7: timing constante. Si el usuario no existe, igual gastamos el tiempo
    # de bcrypt verificando contra un hash señuelo, para que "usuario inexistente"
    # y "contraseña incorrecta" tarden lo mismo y no se puedan enumerar usuarios.
    if not usuario:
        verificar_password(form.password, _HASH_SENUELO)
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Credenciales incorrectas")

    ahora = datetime.now()
    if usuario.bloqueado_hasta and usuario.bloqueado_hasta > ahora:
        faltan = int((usuario.bloqueado_hasta - ahora).total_seconds() / 60) + 1
        raise HTTPException(status_code=423, detail=f"Usuario bloqueado por seguridad. Intente de nuevo en {faltan} minutos")

    if not verificar_password(form.password, usuario.hashed_password):
        usuario.intentos_fallidos += 1
        if usuario.intentos_fallidos >= 3:
            usuario.bloqueado_hasta = ahora + timedelta(minutes=5)
        await db.commit()
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Credenciales incorrectas")

    usuario.intentos_fallidos = 0
    usuario.bloqueado_hasta = None
    await db.commit()

    token = crear_token(data={"sub": usuario.username, "rol": usuario.rol})
    return LoginResponse(access_token=token)


@router.get("/auth/me")
async def usuario_actual(
    admin: Usuario = Depends(obtener_usuario_actual),
):
    """Devuelve los datos del usuario del token. Sirve para validar la sesión
    al recargar el panel (F5) sin pedir login de nuevo."""
    return {
        "username": admin.username,
        "rol": admin.rol,
        "nombre_completo": admin.nombre_completo,
    }


@router.post("/auth/registro", status_code=201)
async def registrar_usuario(
    datos: UsuarioCrear,
    db: AsyncSession = Depends(get_db),
    admin: Usuario = Depends(obtener_usuario_actual)
):
    """Registrar nuevo usuario (solo admins)."""
    if admin.rol != "superadmin":
        raise HTTPException(status_code=403, detail="Solo administradores pueden crear usuarios")
    nuevo = Usuario(
        username=datos.username,
        hashed_password=hashear_password(datos.password),
        nombre_completo=datos.nombre_completo,
        rol=datos.rol
    )
    db.add(nuevo)
    await db.flush()
    await registrar_auditoria(admin.username, "crear_usuario", rol=admin.rol,
                              objetivo=f"{datos.username} (rol={datos.rol})", db=db)
    return {"mensaje": f"Usuario '{datos.username}' creado"}


# ── Alumnos ─────────────────────────────────────────────────────────

@router.post("/alumnos/validar")
async def validar_alumno(request: Request, datos: AlumnoAuth, db: AsyncSession = Depends(get_db)):
    """Validar código/DNI de alumno para desbloquear terminal."""
    # E-6: límite por IP para frenar enumeración masiva de DNIs desde la LAN.
    # 30/min por IP es de sobra para el uso real del kiosco (un login por vez).
    rate_limit.exigir(request, "kiosco", limite=30, ventana_seg=60)
    result = await db.execute(select(AlumnoMaestro).where(
        (AlumnoMaestro.codigo == datos.codigo) | (AlumnoMaestro.dni == datos.codigo)
    ))
    alumno = result.scalar_one_or_none()
    if not alumno:
        raise HTTPException(status_code=404, detail="Alumno no encontrado")
    return {
        "valido": True,
        "codigo": alumno.codigo,
        "nombre": alumno.nombre,
        "escuela": alumno.escuela,
    }


@router.get("/alumnos")
async def listar_alumnos(
    db: AsyncSession = Depends(get_db),
    _: Usuario = Depends(obtener_usuario_actual)
):
    """Listar alumnos del maestro."""
    result = await db.execute(select(AlumnoMaestro).order_by(AlumnoMaestro.nombre))
    return [
        {"dni": a.dni, "codigo": a.codigo, "nombre": a.nombre, "escuela": a.escuela}
        for a in result.scalars().all()
    ]


@router.get("/alumnos/exportar")
async def exportar_alumnos(
    db: AsyncSession = Depends(get_db),
    admin: Usuario = Depends(obtener_usuario_actual)
):
    """Exportar toda la base de datos de alumnos a Excel (Backup)."""
    import io
    from fastapi.responses import StreamingResponse
    import openpyxl
    from sqlalchemy.orm import joinedload

    if admin.rol != "superadmin":
        raise HTTPException(status_code=403, detail="Solo administradores pueden exportar el backup")

    result = await db.execute(
        select(AlumnoMaestro)
        .options(joinedload(AlumnoMaestro.facultad_rel), joinedload(AlumnoMaestro.escuela_rel))
        .order_by(AlumnoMaestro.nombre)
    )
    alumnos = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Alumnos Registrados"
    ws.append(["DNI", "Código", "Nombre Completo", "Facultad", "Escuela"])

    for a in alumnos:
        fac = a.facultad_rel.nombre if a.facultad_rel else ""
        esc = a.escuela_rel.nombre if a.escuela_rel else ""
        ws.append([a.dni, a.codigo, a.nombre, fac, esc])

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    headers = {
        "Content-Disposition": "attachment; filename=backup_alumnos.xlsx",
        "Access-Control-Expose-Headers": "Content-Disposition"
    }
    return StreamingResponse(
        iter([stream.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )


# ── Terminales ──────────────────────────────────────────────────────

@router.get("/terminales")
async def listar_terminales(
    db: AsyncSession = Depends(get_db),
    _: Usuario = Depends(obtener_usuario_actual)
):
    """Listar terminales activas.
    Excluye terminales offline cuya última conexión fue hace más de 2 horas:
    son terminales fantasma que nunca volvieron a conectarse tras un crash.
    Las que están offline pero con conexión reciente (PC recién desconectada)
    sí se incluyen para que el admin las vea unos minutos antes de desaparecer.
    Incluye la sesión activa de cada terminal (si existe) para las tarjetas del panel.
    """
    from datetime import timedelta
    cutoff = datetime.now() - timedelta(hours=2)
    result = await db.execute(select(Terminal).order_by(Terminal.nombre_red))
    terminales = result.scalars().all()

    # Cargar sesiones activas de todas las terminales en una sola consulta
    ids_terminales = [t.id for t in terminales]
    sesiones_activas: dict[int, dict] = {}
    if ids_terminales:
        res_s = await db.execute(
            select(Sesion, AlumnoMaestro)
            .join(AlumnoMaestro, Sesion.dni_alumno == AlumnoMaestro.dni)
            .where(Sesion.id_terminal.in_(ids_terminales), Sesion.estado == "activa")
        )
        for s, a in res_s.all():
            sesiones_activas[s.id_terminal] = {
                "id":           s.id,
                "alumno_nombre": a.nombre,
                "alumno_dni":   a.dni,
                "razon_uso":    s.razon_uso or "",
                "hora_entrada": s.hora_entrada,
                "activa":       True,
                "terminal_ip":  None,   # se rellena abajo
            }

    out = []
    for t in terminales:
        if t.estado == "offline" and not (t.ultima_conexion and t.ultima_conexion >= cutoff):
            continue
        sesion = sesiones_activas.get(t.id)
        if sesion:
            sesion["terminal_ip"] = t.ip
        out.append({
            "id": t.id, "nombre": t.nombre_red, "ip": t.ip,
            "estado": t.estado, "ultima_conexion": t.ultima_conexion,
            "sesion_activa": sesion,
        })
    return out


@router.delete("/terminales/{terminal_id}")
async def eliminar_terminal(
    terminal_id: int,
    db: AsyncSession = Depends(get_db),
    admin: Usuario = Depends(obtener_usuario_actual)
):
    """Eliminar terminal fantasma de la base de datos (solo superadmin)."""
    if admin.rol != "superadmin":
        raise HTTPException(status_code=403, detail="Solo superadmin puede eliminar terminales")
    res = await db.execute(select(Terminal).where(Terminal.id == terminal_id))
    t = res.scalar_one_or_none()
    if not t:
        raise HTTPException(status_code=404, detail="Terminal no encontrada")
    nombre_t = t.nombre_red   # guardar antes del delete (luego el objeto expira)
    await db.delete(t)
    await registrar_auditoria(admin.username, "eliminar_terminal", rol=admin.rol,
                              objetivo=nombre_t, db=db)
    await db.commit()
    return {"ok": True, "eliminada": nombre_t}


@router.post("/terminales/registrar")
async def registrar_terminal(
    datos: TerminalRegistro,
    db: AsyncSession = Depends(get_db),
    _: Usuario = Depends(obtener_usuario_actual)
):
    """Registrar o actualizar terminal al conectarse."""
    result = await db.execute(select(Terminal).where(Terminal.ip == datos.ip))
    terminal = result.scalar_one_or_none()
    if terminal:
        terminal.nombre = datos.nombre
        terminal.ultima_conexion = datetime.now()
        terminal.estado = "bloqueado"
    else:
        terminal = Terminal(nombre_red=datos.nombre, ip=datos.ip, estado="bloqueado", ultima_conexion=datetime.now())
        db.add(terminal)
    await db.flush()
    return {"id": terminal.id, "nombre": terminal.nombre}


# ── Sesiones ────────────────────────────────────────────────────────

@router.post("/sesiones/iniciar")
async def iniciar_sesion(
    alumno_codigo: str,
    terminal_ip: str,
    db: AsyncSession = Depends(get_db),
    _: Usuario = Depends(obtener_usuario_actual),   # F-4: exigir auth (antes era público)
):
    """Iniciar sesión de uso en una terminal."""
    res_alumno = await db.execute(select(AlumnoMaestro).where(
        (AlumnoMaestro.codigo == alumno_codigo) | (AlumnoMaestro.dni == alumno_codigo)
    ))
    alumno = res_alumno.scalar_one_or_none()
    if not alumno:
        raise HTTPException(status_code=403, detail="Alumno no encontrado")

    res_terminal = await db.execute(select(Terminal).where(Terminal.ip == terminal_ip))
    terminal = res_terminal.scalar_one_or_none()
    if not terminal:
        raise HTTPException(status_code=404, detail="Terminal no registrada")

    sesion = Sesion(
        dni_alumno=alumno.dni,
        id_terminal=terminal.id,
        hora_entrada=datetime.now(),
        fecha_uso=datetime.now().date(),
    )
    terminal.estado = "activo"
    db.add(sesion)
    await db.flush()

    # ── Detección de anomalías ────────────────────────────────────
    ahora = datetime.now()

    # 1. Cambios rápidos de PC: 3+ sesiones en terminales distintas en los últimos 5 min
    ventana = ahora - timedelta(minutes=5)
    res_rec = await db.execute(
        select(Sesion)
        .where(Sesion.dni_alumno == alumno.dni, Sesion.hora_entrada >= ventana)
        .order_by(Sesion.hora_entrada.desc())
    )
    sesiones_recientes = res_rec.scalars().all()
    terminales_distintas = {s.id_terminal for s in sesiones_recientes}
    if len(terminales_distintas) >= 3:
        detalle = (
            f"{alumno.nombre} inició sesión en {len(terminales_distintas)} PCs distintas "
            f"en los últimos 5 minutos ({len(sesiones_recientes)} intentos). "
            f"Posible uso compartido de credencial."
        )
        sospecha = Sospecha(
            dni_alumno=alumno.dni,
            nombre_alumno=alumno.nombre,
            tipo="cambio_pc_rapido",
            detalle=detalle,
        )
        db.add(sospecha)
        await db.flush()
        await manager.broadcast({
            "tipo": "sospecha",
            "nivel": "alerta",
            "mensaje": f"SOSPECHA - {alumno.nombre} (DNI {alumno.dni}): cambió de PC "
                       f"{len(terminales_distintas)} veces en 5 min",
            "sospecha_id": sospecha.id,
        })

    await db.commit()
    return {"sesion_id": sesion.id, "mensaje": "Sesión iniciada", "alumno": alumno.nombre}


@router.post("/sesiones/{sesion_id}/cerrar")
async def cerrar_sesion(
    sesion_id: int,
    motivo: str = "manual",
    db: AsyncSession = Depends(get_db),
    _: Usuario = Depends(obtener_usuario_actual),   # F-4: exigir auth (el panel ya manda el token)
):
    """Cerrar una sesión activa usando exclusivamente el reloj del servidor.

    Idempotente: si la sesión ya está cerrada (porque la PC se desconectó/apagó
    antes de que llegara este POST, o fue cancelada como fantasma), devolvemos
    éxito en vez de 404 — el objetivo (sesión cerrada) ya se cumplió. Solo es
    404 si la sesión nunca existió.
    """
    result = await db.execute(select(Sesion).where(Sesion.id == sesion_id))
    sesion = result.scalar_one_or_none()
    if not sesion:
        raise HTTPException(status_code=404, detail="Sesión no encontrada")

    # Ya cerrada por otra vía (desconexión, fantasma, logout) — nada que hacer.
    if sesion.estado != "activa":
        # Aseguramos que la terminal quede bloqueada de todos modos.
        await db.execute(
            update(Terminal).where(Terminal.id == sesion.id_terminal).values(estado="bloqueado")
        )
        await db.commit()
        return {
            "mensaje": "Sesión ya estaba cerrada",
            "ya_cerrada": True,
            "motivo_previo": sesion.motivo_cierre,
        }

    ahora = datetime.now().replace(tzinfo=None)
    sesion.hora_salida   = ahora
    sesion.activa        = False
    sesion.motivo_cierre = motivo

    await db.execute(
        update(Terminal).where(Terminal.id == sesion.id_terminal).values(estado="bloqueado")
    )
    inicio_naive = sesion.inicio.replace(tzinfo=None) if sesion.inicio else ahora
    duracion_min = int((ahora - inicio_naive).total_seconds() / 60)

    await db.commit()
    return {"mensaje": "Sesión cerrada", "duracion_min": duracion_min, "hora_salida": ahora.strftime("%I:%M:%S %p")}


@router.get("/catalogos/motivos")
async def obtener_motivos_activos(db: AsyncSession = Depends(get_db)):
    """Devuelve la lista de motivos de uso activos para el kiosco."""
    from models import CatalogoMotivo
    result = await db.execute(select(CatalogoMotivo).where(CatalogoMotivo.activo == True))
    motivos = result.scalars().all()
    return [{"id": m.id, "descripcion": m.descripcion} for m in motivos]


# Mapeo columna -> campos ORM para ORDER BY.
# "escuela" y "estudiante" usan Alumno para agrupación real con JOIN.
_SORT_MAP = {
    "estudiante": (AlumnoMaestro.nombre,),
    "codigo":     (AlumnoMaestro.codigo,),
    "dni":        (AlumnoMaestro.dni,),
    "facultad":   (Sesion.razon_uso,),   # fallback; facultad se resuelve por JOIN
    "escuela":    (AlumnoMaestro.nombre,),
    "actividad":  (Sesion.razon_uso,),
    "fecha":      (Sesion.hora_entrada,),
    "inicio":     (Sesion.hora_entrada,),
}


def _aplicar_filtro_fecha(q, periodo: str | None, fecha_inicio: str | None, fecha_fin: str | None,
                          mes: str | None = None, anio: str | None = None):
    """Aplica filtro de fecha sobre Sesion.fecha_uso (columna DATE en MySQL)."""
    from datetime import date
    import calendar

    hoy = date.today()

    # Mes + año específico tiene prioridad sobre los chips rápidos
    if mes and anio:
        anio_i = int(anio)
        mes_i  = int(mes)
        ultimo = calendar.monthrange(anio_i, mes_i)[1]
        q = q.where(
            Sesion.fecha_uso >= date(anio_i, mes_i, 1),
            Sesion.fecha_uso <= date(anio_i, mes_i, ultimo),
        )
    elif anio and not mes:
        anio_i = int(anio)
        q = q.where(
            Sesion.fecha_uso >= date(anio_i, 1, 1),
            Sesion.fecha_uso <= date(anio_i, 12, 31),
        )
    elif mes and not anio:
        mes_i  = int(mes)
        anio_i = hoy.year
        ultimo = calendar.monthrange(anio_i, mes_i)[1]
        q = q.where(
            Sesion.fecha_uso >= date(anio_i, mes_i, 1),
            Sesion.fecha_uso <= date(anio_i, mes_i, ultimo),
        )
    elif periodo == "dia" or periodo is None:
        q = q.where(Sesion.fecha_uso == hoy)
    elif periodo == "mes":
        import calendar as _cal
        ultimo_dia = _cal.monthrange(hoy.year, hoy.month)[1]
        q = q.where(
            Sesion.fecha_uso >= hoy.replace(day=1),
            Sesion.fecha_uso <= hoy.replace(day=ultimo_dia),
        )
    elif periodo == "anio":
        q = q.where(
            Sesion.fecha_uso >= hoy.replace(month=1, day=1),
            Sesion.fecha_uso <= hoy.replace(month=12, day=31),
        )
    elif periodo == "rango" and fecha_inicio and fecha_fin:
        from datetime import datetime as _dt
        q = q.where(
            Sesion.fecha_uso >= _dt.strptime(fecha_inicio, "%Y-%m-%d").date(),
            Sesion.fecha_uso <= _dt.strptime(fecha_fin,    "%Y-%m-%d").date(),
        )
    # Si periodo == "todo" no aplica filtro
    return q


@router.get("/sesiones/activas")
async def sesiones_activas(
    db: AsyncSession = Depends(get_db),
    _: Usuario = Depends(obtener_usuario_actual),
    search:       str | None = None,
    actividad:    str | None = None,
    sort_by:      str = "fecha",
    order:        str = "desc",
    periodo:      str | None = None,
    fecha_inicio: str | None = None,
    fecha_fin:    str | None = None,
    mes:          str | None = None,
    anio:         str | None = None,
    limit:        int = 30,
    offset:       int = 0,
):
    """Historial con filtros de búsqueda, actividad, fecha y ordenamiento."""
    from sqlalchemy import or_

    from sqlalchemy.orm import aliased
    FacDir = aliased(Facultad, name="fac_direct")

    q = (select(Sesion, AlumnoMaestro, Terminal, Escuela, Facultad, FacDir)
         .join(AlumnoMaestro, Sesion.dni_alumno == AlumnoMaestro.dni)
         .join(Terminal, Sesion.id_terminal == Terminal.id)
         .outerjoin(Escuela,  AlumnoMaestro.id_escuela  == Escuela.id)
         .outerjoin(Facultad, Escuela.id_facultad        == Facultad.id)
         .outerjoin(FacDir,   AlumnoMaestro.id_facultad  == FacDir.id))

    q = _aplicar_filtro_fecha(q, periodo, fecha_inicio, fecha_fin, mes, anio)

    if search:
        like = f"%{search}%"
        q = q.where(or_(
            AlumnoMaestro.nombre.ilike(like),
            AlumnoMaestro.dni.ilike(like),
            AlumnoMaestro.codigo.ilike(like),
        ))
    if actividad:
        q = q.where(Sesion.razon_uso.ilike(f"%{actividad}%"))

    cols = _SORT_MAP.get(sort_by.lower(), (Sesion.hora_entrada,))
    q = q.order_by(*cols) if order.lower() == "asc" else q.order_by(*[c.desc() for c in cols])

    from sqlalchemy import func as _func
    total_result = await db.execute(select(_func.count()).select_from(q.subquery()))
    total = total_result.scalar() or 0

    limit = min(max(limit, 1), 200)   # G-13: tope de seguridad
    q = q.offset(offset).limit(limit)
    result = await db.execute(q)
    rows = []
    for s, a, t, esc_obj, fac_obj, fac_dir in result.all():
        inicio_naive = s.hora_entrada.replace(tzinfo=None) if s.hora_entrada else None
        salida_naive = s.hora_salida.replace(tzinfo=None)  if s.hora_salida  else None
        duracion_min = int((salida_naive - inicio_naive).total_seconds() / 60) if (inicio_naive and salida_naive) else None
        fac = (fac_obj.nombre if fac_obj else None) or (fac_dir.nombre if fac_dir else "")
        esc = esc_obj.nombre if esc_obj else ""
        rows.append({
            "id":              s.id,
            "inicio":          s.hora_entrada,
            "hora_salida":     s.hora_salida,
            "hora_salida_fmt": salida_naive.strftime("%I:%M:%S %p") if salida_naive else None,
            "fin":             s.hora_salida,
            "fecha_uso":       s.fecha_uso or (s.hora_entrada.date() if s.hora_entrada else None),
            "alumno_codigo":   a.codigo or "",
            "alumno_dni":      a.dni,
            "alumno_nombre":   a.nombre,
            "dni":             a.dni,
            "facultad":        fac,
            "escuela":         esc,
            "terminal_nombre": t.nombre_red,
            "terminal_ip":     t.ip,
            "razon_uso":       s.razon_uso or "",
            "activa":          s.activa,
            "duracion_min":    duracion_min,
        })
    return {"total": total, "items": rows}


def _resolver_mysqldump() -> str | None:
    """Localiza el ejecutable mysqldump de forma robusta, sin depender del PATH
    del momento en que arrancó el servidor (que puede no incluir MySQL).

    Orden de búsqueda:
      1. Variable MYSQLDUMP_PATH del .env (ruta absoluta explícita).
      2. shutil.which('mysqldump') — el PATH actual del proceso.
      3. Carpetas típicas de instalación de MySQL en Windows
         (Program Files\\MySQL\\MySQL Server X.Y\\bin), tomando la más nueva.
    Devuelve la ruta absoluta o None si no se encuentra.
    """
    import os, shutil, glob

    # 1. Ruta explícita en .env
    ruta_env = os.getenv("MYSQLDUMP_PATH", "").strip()
    if ruta_env and os.path.isfile(ruta_env):
        return ruta_env

    # 2. PATH del proceso
    encontrado = shutil.which("mysqldump")
    if encontrado:
        return encontrado

    # 3. Carpetas típicas de Windows. Se respeta el ORDEN de los patrones
    #    (MySQL Server real primero; XAMPP como último recurso porque su
    #    mysqldump suele ser incompatible con caching_sha2_password). Dentro de
    #    cada patrón se prefiere la versión más nueva.
    patrones = [
        r"C:\Program Files\MySQL\MySQL Server*\bin\mysqldump.exe",
        r"C:\Program Files (x86)\MySQL\MySQL Server*\bin\mysqldump.exe",
        r"C:\xampp\mysql\bin\mysqldump.exe",  # último recurso
    ]
    for patron in patrones:
        matches = glob.glob(patron)
        if matches:
            matches.sort(reverse=True)  # versión más nueva dentro del patrón
            return matches[0]

    return None


@router.get("/admin/backup-sql")
async def backup_sql(
    admin: Usuario = Depends(obtener_usuario_actual),
):
    """Genera un volcado SQL completo de la BD y lo devuelve como descarga."""
    import re
    from fastapi.responses import StreamingResponse
    from datetime import datetime as _dt
    import os

    # A-7: solo superadmin puede exportar la base completa (contiene PII de
    # todos los alumnos, sesiones e incidencias). Un admin básico no.
    if admin.rol != "superadmin":
        raise HTTPException(status_code=403, detail="Solo superadmin puede exportar la base de datos")

    # Parsear credenciales desde DATABASE_URL
    db_url = os.getenv("DATABASE_URL", "")
    m = re.match(r"mysql\+asyncmy://([^:]+):([^@]+)@([^:/]+):?(\d+)?/(.+)", db_url)
    if not m:
        raise HTTPException(status_code=500, detail="No se pudo leer la configuración de BD")

    user, password, host, port, dbname = m.group(1), m.group(2), m.group(3), m.group(4) or "3306", m.group(5)
    timestamp = _dt.now().strftime("%Y%m%d_%H%M%S")
    filename  = f"backup_{dbname}_{timestamp}.sql"

    # Localizar mysqldump sin depender del PATH heredado al arrancar el servidor.
    mysqldump_bin = _resolver_mysqldump()
    if not mysqldump_bin:
        raise HTTPException(
            status_code=500,
            detail="No se encontró 'mysqldump'. Instale MySQL o defina MYSQLDUMP_PATH en .env.",
        )

    # A-6: NO pasar la contraseña como argumento (--password=) porque queda
    # visible en la lista de procesos (tasklist/ps). Se pasa por la variable
    # de entorno MYSQL_PWD, que mysqldump lee automáticamente y no se expone
    # en la línea de comando.
    cmd = [
        mysqldump_bin,
        f"--user={user}",
        f"--host={host}",
        f"--port={port}",
        "--single-transaction",   # consistencia sin bloquear tablas InnoDB
        "--routines",             # incluir stored procedures/functions
        "--triggers",             # incluir triggers
        "--no-tablespaces",       # evita error de permisos en MariaDB/MySQL 5.7+
        "--complete-insert",      # INSERT INTO con nombres de columna explícitos
        "--extended-insert",      # agrupa filas en un solo INSERT (más eficiente)
        "--add-drop-table",       # DROP TABLE IF EXISTS antes de cada CREATE
        "--disable-keys",         # deshabilita índices durante restore (más rápido)
        "--hex-blob",             # exporta BLOB/BINARY en hex para evitar corrupción
        dbname,
    ]
    env = {**os.environ, "MYSQL_PWD": password}

    proc = await asyncio.create_subprocess_exec(
        *cmd,
        stdout=asyncio.subprocess.PIPE,
        stderr=asyncio.subprocess.PIPE,
        env=env,
    )
    stdout, stderr = await proc.communicate()

    if proc.returncode != 0:
        from main import logger
        logger.error(f"[BACKUP] mysqldump falló: {stderr.decode()[:500]}")
        raise HTTPException(status_code=500, detail="No se pudo generar el backup. Revise los logs del servidor.")

    sql_text = stdout.decode("utf-8", errors="replace")

    # Verificar que el dump incluye datos reales (al menos un INSERT)
    from main import logger
    if "INSERT INTO" not in sql_text:
        logger.warning("[BACKUP] El dump no contiene INSERTs — puede que la BD esté vacía o mysqldump falló silenciosamente")

    logger.info(f"[BACKUP] Generado: {filename} ({len(stdout):,} bytes)")

    # Auditoría en sesión propia (este endpoint no abre transacción de escritura)
    await registrar_auditoria(admin.username, "backup_sql", rol=admin.rol,
                              objetivo=dbname, detalle=f"{len(stdout):,} bytes")

    return StreamingResponse(
        iter([stdout]),
        media_type="application/octet-stream",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Sospechas ────────────────────────────────────────────────────────

@router.get("/admin/sospechas")
async def listar_sospechas(
    estado: str | None = None,   # pendiente | aprobada | descartada
    limit: int = 100,
    offset: int = 0,
    db: AsyncSession = Depends(get_db),
    _: Usuario = Depends(obtener_usuario_actual),
):
    from sqlalchemy import func as _func
    q = select(Sospecha).order_by(Sospecha.fecha.desc())
    if estado:
        q = q.where(Sospecha.estado == estado)
    limit = min(max(limit, 1), 200)   # G-13: tope de seguridad
    total = (await db.execute(select(_func.count()).select_from(q.subquery()))).scalar()
    rows  = (await db.execute(q.offset(offset).limit(limit))).scalars().all()
    return {
        "total": total,
        "items": [
            {
                "id":            s.id,
                "dni":           s.dni_alumno,
                "nombre_alumno": s.nombre_alumno,
                "tipo":          s.tipo,
                "detalle":       s.detalle,
                "fecha":         s.fecha,
                "estado":        s.estado,
                "revisado_por":  s.revisado_por,
                "fecha_revision": s.fecha_revision,
            }
            for s in rows
        ],
    }


@router.post("/admin/sospechas/{sospecha_id}/aprobar")
async def aprobar_sospecha(
    sospecha_id: int,
    db: AsyncSession = Depends(get_db),
    admin: Usuario = Depends(obtener_usuario_actual),
):
    """Convierte la sospecha en incidencia leve y la marca como aprobada."""
    res = await db.execute(select(Sospecha).where(Sospecha.id == sospecha_id))
    sospecha = res.scalar_one_or_none()
    if not sospecha:
        raise HTTPException(status_code=404, detail="Sospecha no encontrada")
    if sospecha.estado != "pendiente":
        raise HTTPException(status_code=409, detail="La sospecha ya fue revisada")

    inc = Incidencia(
        dni_alumno=sospecha.dni_alumno,
        nombre_alumno=sospecha.nombre_alumno,
        tipo="leve",
        motivo=_TIPO_LABEL.get(sospecha.tipo, sospecha.tipo),
        descripcion=sospecha.detalle,
        registrado_por=f"sistema (aprobado por {admin.username})",
    )
    db.add(inc)

    sospecha.estado        = "aprobada"
    sospecha.revisado_por  = admin.username
    sospecha.fecha_revision = datetime.now()
    await registrar_auditoria(admin.username, "aprobar_sospecha", rol=admin.rol,
                              objetivo=f"DNI {sospecha.dni_alumno}",
                              detalle="creó incidencia leve", db=db)
    await db.commit()
    return {"mensaje": "Sospecha aprobada e incidencia registrada"}


@router.post("/admin/sospechas/{sospecha_id}/descartar")
async def descartar_sospecha(
    sospecha_id: int,
    db: AsyncSession = Depends(get_db),
    admin: Usuario = Depends(obtener_usuario_actual),
):
    res = await db.execute(select(Sospecha).where(Sospecha.id == sospecha_id))
    sospecha = res.scalar_one_or_none()
    if not sospecha:
        raise HTTPException(status_code=404, detail="Sospecha no encontrada")
    if sospecha.estado != "pendiente":
        raise HTTPException(status_code=409, detail="La sospecha ya fue revisada")

    sospecha.estado         = "descartada"
    sospecha.revisado_por   = admin.username
    sospecha.fecha_revision = datetime.now()
    await registrar_auditoria(admin.username, "descartar_sospecha", rol=admin.rol,
                              objetivo=f"DNI {sospecha.dni_alumno}", db=db)
    await db.commit()
    return {"mensaje": "Sospecha descartada"}


_TIPO_LABEL = {
    "cambio_pc_rapido":   "Cambio rápido de PC (posible uso compartido)",
    "dni_baneado_intento": "Intento de ingreso con DNI baneado",
    "sesion_larga":        "Sesión excesivamente larga sin cierre",
}


# ── Actividad Logs ───────────────────────────────────────────────────

@router.get("/admin/actividad")
async def listar_actividad(
    dni:      str | None = None,
    terminal: str | None = None,
    nivel:    str | None = None,   # normal | sospechoso
    fecha:    str | None = None,   # YYYY-MM-DD
    incluir_ignorados: bool = False,  # True = mostrar también procesos ocultados
    limit:    int = 200,
    offset:   int = 0,
    db: AsyncSession = Depends(get_db),
    _: Usuario = Depends(obtener_usuario_actual),
):
    from sqlalchemy import func as _func
    q = select(ActividadLog).order_by(ActividadLog.fecha_hora.desc())
    if dni:
        q = q.where(ActividadLog.dni_alumno == dni)
    if terminal:
        q = q.where(ActividadLog.nombre_terminal.ilike(f"%{terminal}%"))
    if nivel:
        q = q.where(ActividadLog.nivel == nivel)

    # Por defecto se ocultan los procesos marcados como ignorados (vista limpia).
    # Los sospechosos SIEMPRE se muestran, aunque su exe esté en la lista.
    # incluir_ignorados=True (al investigar una sospecha) muestra todo.
    if not incluir_ignorados:
        subq_ign = select(ProcesoIgnorado.nombre_exe)
        q = q.where(
            (ActividadLog.nivel == "sospechoso")
            | (ActividadLog.proceso_exe.is_(None))
            | (ActividadLog.proceso_exe.notin_(subq_ign))
        )
    if fecha:
        try:
            d = date.fromisoformat(fecha)
            # BUG-5: usar < inicio del día siguiente para incluir el día completo
            # (antes < 23:59:59 dejaba fuera el último segundo del día).
            q = q.where(ActividadLog.fecha_hora >= datetime(d.year, d.month, d.day))
            q = q.where(ActividadLog.fecha_hora <  datetime(d.year, d.month, d.day) + timedelta(days=1))
        except ValueError:
            pass

    limit = min(max(limit, 1), 200)   # G-13: tope de seguridad
    total = (await db.execute(select(_func.count()).select_from(q.subquery()))).scalar()
    rows  = (await db.execute(q.offset(offset).limit(limit))).scalars().all()

    return {
        "total": total,
        "items": [
            {
                "id":              r.id,
                "id_terminal":     r.id_terminal,
                "nombre_terminal": r.nombre_terminal,
                "dni_alumno":      r.dni_alumno,
                "nombre_alumno":   r.nombre_alumno,
                "tipo":            r.tipo,
                "descripcion":     r.descripcion,
                "detalle":         r.detalle,
                "proceso_exe":     r.proceso_exe,
                "nivel":           r.nivel,
                "fecha_hora":      r.fecha_hora,
            }
            for r in rows
        ],
    }


@router.get("/admin/actividad/resumen-hoy")
async def resumen_actividad_hoy(
    db: AsyncSession = Depends(get_db),
    _: Usuario = Depends(obtener_usuario_actual),
):
    """Conteo de eventos sospechosos del día para el badge del menú."""
    from sqlalchemy import func as _func
    hoy = date.today()
    desde = datetime(hoy.year, hoy.month, hoy.day, 0, 0, 0)
    total_sosp = (await db.execute(
        select(_func.count()).where(
            ActividadLog.nivel == "sospechoso",
            ActividadLog.fecha_hora >= desde,
        )
    )).scalar()
    total_hoy = (await db.execute(
        select(_func.count()).where(ActividadLog.fecha_hora >= desde)
    )).scalar()
    return {"sospechosos_hoy": total_sosp, "total_hoy": total_hoy}


@router.get("/admin/actividad/resumen-por-pc")
async def resumen_actividad_por_pc(
    db: AsyncSession = Depends(get_db),
    _: Usuario = Depends(obtener_usuario_actual),
):
    """Resumen de actividad del día agrupado por PC, para las tarjetas del panel."""
    from sqlalchemy import func as _func, case
    hoy   = date.today()
    desde = datetime(hoy.year, hoy.month, hoy.day, 0, 0, 0)

    # Total eventos y sospechosos por PC hoy
    q = (
        select(
            ActividadLog.nombre_terminal,
            ActividadLog.id_terminal,
            _func.count().label("total"),
            _func.sum(case((ActividadLog.nivel == "sospechoso", 1), else_=0)).label("sospechosos"),
            _func.max(ActividadLog.fecha_hora).label("ultimo_evento"),
            _func.max(ActividadLog.descripcion).label("ultima_descripcion"),
        )
        .where(ActividadLog.fecha_hora >= desde)
        .group_by(ActividadLog.nombre_terminal, ActividadLog.id_terminal)
        .order_by(_func.max(ActividadLog.fecha_hora).desc())
    )
    rows = (await db.execute(q)).all()
    return {
        "items": [
            {
                "nombre_terminal":    r.nombre_terminal,
                "id_terminal":        r.id_terminal,
                "total_hoy":          r.total,
                "sospechosos_hoy":    int(r.sospechosos or 0),
                "ultimo_evento":      r.ultimo_evento,
                "ultima_descripcion": r.ultima_descripcion,
            }
            for r in rows
        ]
    }


# ── Procesos ignorados (lista negra editable) ───────────────────────

@router.get("/admin/procesos-ignorados")
async def listar_procesos_ignorados(
    db: AsyncSession = Depends(get_db),
    _: Usuario = Depends(obtener_usuario_actual),
):
    """Lista los procesos que el admin marcó para no registrar como actividad."""
    rows = (await db.execute(
        select(ProcesoIgnorado).order_by(ProcesoIgnorado.nombre_exe)
    )).scalars().all()
    return {
        "items": [
            {"id": r.id, "nombre_exe": r.nombre_exe,
             "agregado_por": r.agregado_por, "fecha": r.fecha}
            for r in rows
        ]
    }


@router.post("/admin/procesos-ignorados")
async def agregar_proceso_ignorado(
    nombre_exe: str = Body(..., embed=True),
    db: AsyncSession = Depends(get_db),
    admin: Usuario = Depends(obtener_usuario_actual),
):
    """Agrega un proceso a la lista de ignorados. Aplica al instante a todas las PCs."""
    nombre_exe = (nombre_exe or "").strip()
    if not nombre_exe:
        raise HTTPException(status_code=422, detail="El nombre del ejecutable no puede estar vacío")

    # Evitar duplicados (idempotente).
    existe = (await db.execute(
        select(ProcesoIgnorado).where(ProcesoIgnorado.nombre_exe == nombre_exe)
    )).scalar_one_or_none()
    if existe:
        return {"ok": True, "ya_existia": True, "nombre_exe": nombre_exe}

    pi = ProcesoIgnorado(nombre_exe=nombre_exe, agregado_por=admin.username)
    db.add(pi)
    await registrar_auditoria(admin.username, "agregar_proceso_ignorado", rol=admin.rol,
                              objetivo=nombre_exe, db=db)
    await db.commit()
    return {"ok": True, "id": pi.id, "nombre_exe": nombre_exe}


@router.delete("/admin/procesos-ignorados/{proceso_id}")
async def eliminar_proceso_ignorado(
    proceso_id: int,
    db: AsyncSession = Depends(get_db),
    admin: Usuario = Depends(obtener_usuario_actual),
):
    """Quita un proceso de la lista de ignorados — volverá a registrarse."""
    pi = (await db.execute(
        select(ProcesoIgnorado).where(ProcesoIgnorado.id == proceso_id)
    )).scalar_one_or_none()
    if not pi:
        raise HTTPException(status_code=404, detail="Proceso no encontrado")
    nombre = pi.nombre_exe
    await db.delete(pi)
    await registrar_auditoria(admin.username, "quitar_proceso_ignorado", rol=admin.rol,
                              objetivo=nombre, db=db)
    await db.commit()
    return {"ok": True, "eliminado": nombre}


@router.get("/admin/exportar-excel")
async def exportar_excel(
    db: AsyncSession = Depends(get_db),
    _: Usuario = Depends(obtener_usuario_actual)
):
    """Genera y descarga un archivo Excel con el historial completo de sesiones."""
    import io
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    result = await db.execute(
        select(Sesion, AlumnoMaestro, Terminal, Escuela, Facultad)
        .join(AlumnoMaestro, Sesion.dni_alumno == AlumnoMaestro.dni)
        .join(Terminal, Sesion.id_terminal == Terminal.id)
        .outerjoin(Escuela, AlumnoMaestro.id_escuela == Escuela.id)
        .outerjoin(Facultad, Escuela.id_facultad == Facultad.id)
        .order_by(Sesion.hora_entrada.desc())
    )
    rows = result.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Historial de Sesiones"

    headers = ["Estudiante", "Código", "DNI", "Facultad", "Escuela", "Actividad", "Equipo/PC", "Inicio", "Salida", "Fecha", "Duración (min)"]
    header_fill = PatternFill("solid", fgColor="1E40AF")
    header_font = Font(color="FFFFFF", bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, (s, a, t, esc, fac) in enumerate(rows, 2):
        inicio_naive = s.hora_entrada.replace(tzinfo=None) if s.hora_entrada else None
        fin_naive    = s.hora_salida.replace(tzinfo=None)  if s.hora_salida  else None
        duracion     = int((fin_naive - inicio_naive).total_seconds() / 60) if (inicio_naive and fin_naive) else ""
        ws.append([
            a.nombre,
            a.codigo or "",
            a.dni,
            fac.nombre  if fac else "",
            esc.nombre  if esc else "",
            s.razon_uso or "",
            t.nombre_red,
            inicio_naive.strftime("%I:%M:%S %p") if inicio_naive else "",
            fin_naive.strftime("%I:%M:%S %p")    if fin_naive    else "En curso",
            inicio_naive.strftime("%d/%m/%Y")    if inicio_naive else "",
            duracion,
        ])

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"historial_sesiones_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/admin/exportar-pdf")
async def exportar_pdf(
    db: AsyncSession = Depends(get_db),
    _: Usuario = Depends(obtener_usuario_actual),
    search:       str | None = None,
    actividad:    str | None = None,
    sort_by:      str = "fecha",
    order:        str = "desc",
    periodo:      str | None = None,
    fecha_inicio: str | None = None,
    fecha_fin:    str | None = None,
    mes:          str | None = None,
    anio:         str | None = None,
):
    """Genera PDF con estadísticas rápidas + tabla filtrada/ordenada."""
    import io
    from collections import Counter
    from fastapi.responses import StreamingResponse
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    from sqlalchemy import or_

    q = (select(Sesion, AlumnoMaestro, Terminal, Escuela, Facultad)
         .join(AlumnoMaestro, Sesion.dni_alumno == AlumnoMaestro.dni)
         .join(Terminal, Sesion.id_terminal == Terminal.id)
         .outerjoin(Escuela, AlumnoMaestro.id_escuela == Escuela.id)
         .outerjoin(Facultad, Escuela.id_facultad == Facultad.id))
    q = _aplicar_filtro_fecha(q, periodo, fecha_inicio, fecha_fin, mes, anio)
    if search:
        like = f"%{search}%"
        q = q.where(or_(
            AlumnoMaestro.nombre.ilike(like),
            AlumnoMaestro.dni.ilike(like),
            AlumnoMaestro.codigo.ilike(like),
        ))
    if actividad:
        q = q.where(Sesion.razon_uso.ilike(f"%{actividad}%"))

    cols = _SORT_MAP.get(sort_by.lower(), (Sesion.hora_entrada,))
    q = q.order_by(*cols) if order.lower() == "asc" else q.order_by(*[c.desc() for c in cols])

    result = await db.execute(q)
    rows = result.all()

    # ── Calcular estadísticas ─────────────────────────────────────────
    alumnos_unicos  = len({a.dni for _, a, _, _, _ in rows})
    razones         = [s.razon_uso for s, _, _, _, _ in rows if s.razon_uso]
    actividad_top   = Counter(razones).most_common(1)[0][0] if razones else "—"
    total_min       = 0
    for s, _, _, _, _ in rows:
        ini = s.hora_entrada.replace(tzinfo=None) if s.hora_entrada else None
        sal = s.hora_salida.replace(tzinfo=None)  if s.hora_salida  else None
        if ini and sal:
            total_min += int((sal - ini).total_seconds() / 60)
    horas, mins = divmod(total_min, 60)

    # ── Estilos ───────────────────────────────────────────────────────
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1*cm, rightMargin=1*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles  = getSampleStyleSheet()
    c_azul  = colors.HexColor("#1E40AF")
    c_claro = colors.HexColor("#EFF6FF")
    c_amber = colors.HexColor("#FEF3C7")
    c_amber_brd = colors.HexColor("#F59E0B")
    title_st = ParagraphStyle("titulo", parent=styles["Title"],
                              fontSize=14, textColor=c_azul, alignment=TA_CENTER)
    sub_st   = ParagraphStyle("sub", parent=styles["Normal"],
                              fontSize=8, textColor=colors.grey, alignment=TA_CENTER)
    filtro_st = ParagraphStyle("filtro", parent=styles["Normal"],
                               fontSize=9, textColor=colors.HexColor("#92400E"),
                               alignment=TA_CENTER, leading=13)

    # Construir descripción de filtros activos
    _periodos = {"dia": "Hoy", "mes": "Este mes", "anio": "Este año",
                 "rango": f"{fecha_inicio} a {fecha_fin}", "todo": "Todo el historial"}
    periodo_desc = _periodos.get(periodo or "dia", "Hoy")

    elementos = [
        Paragraph("Historial de Sesiones — Biblioteca UNASAM", title_st),
        Paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %I:%M:%S %p')}", sub_st),
        Spacer(1, 0.2*cm),
    ]

    # Bloque de filtros activos (solo si hay algo distinto al predeterminado)
    filtros_activos = []
    if periodo == "rango" and fecha_inicio and fecha_fin:
        filtros_activos.append(f"Período: {fecha_inicio} a {fecha_fin}")
    elif periodo and periodo != "dia":
        filtros_activos.append(f"Período: {periodo_desc}")
    if search:
        filtros_activos.append(f"Búsqueda: «{search}»")
    if actividad:
        filtros_activos.append(f"Actividad: «{actividad}»")

    if filtros_activos:
        filtro_tabla = Table(
            [[Paragraph("  ".join(filtros_activos), filtro_st)]],
            colWidths=[doc.width]
        )
        filtro_tabla.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, -1), c_amber),
            ("BOX",           (0, 0), (-1, -1), 1, c_amber_brd),
            ("TOPPADDING",    (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("LEFTPADDING",   (0, 0), (-1, -1), 8),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 8),
        ]))
        elementos += [filtro_tabla, Spacer(1, 0.25*cm)]
    else:
        elementos.append(Spacer(1, 0.15*cm))


    # ── Cuadro de estadísticas ────────────────────────────────────────
    stats_data = [
        ["Total sesiones", "Alumnos únicos", "Actividad frecuente", "Tiempo total de uso"],
        [str(len(rows)), str(alumnos_unicos), actividad_top, f"{horas}h {mins}min"],
    ]
    stats_table = Table(stats_data, colWidths=[6*cm, 5*cm, 9*cm, 5.5*cm])
    stats_table.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0),  c_azul),
        ("TEXTCOLOR",     (0, 0), (-1, 0),  colors.white),
        ("FONTNAME",      (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, 0),  8),
        ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME",      (0, 1), (-1, 1),  "Helvetica-Bold"),
        ("FONTSIZE",      (0, 1), (-1, 1),  12),
        ("TEXTCOLOR",     (0, 1), (-1, 1),  c_azul),
        ("BACKGROUND",    (0, 1), (-1, 1),  c_claro),
        ("BOX",           (0, 0), (-1, -1), 1,   c_azul),
        ("INNERGRID",     (0, 0), (-1, -1), 0.5, colors.HexColor("#93C5FD")),
        ("TOPPADDING",    (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    elementos += [stats_table, Spacer(1, 0.4*cm)]

    # ── Tabla principal ───────────────────────────────────────────────
    # Estilos de párrafo para word-wrap en celdas
    cell_st = ParagraphStyle("cell", parent=styles["Normal"],
                             fontSize=6.5, leading=9, wordWrap="LTR")
    fac_st  = ParagraphStyle("fac", parent=styles["Normal"],
                             fontSize=6.0, leading=7.5, wordWrap="LTR") # Ligeramente menor para textos largos
    hdr_st  = ParagraphStyle("hdr",  parent=styles["Normal"],
                             fontSize=7, fontName="Helvetica-Bold",
                             textColor=colors.white, alignment=TA_CENTER, leading=9)

    headers_p  = [Paragraph(h, hdr_st) for h in
                  ["Estudiante", "Código", "DNI", "Facultad", "Escuela", "Actividad", "Equipo", "Inicio", "Salida", "Fecha", "Min"]]
    # Ajustar anchos para que quepa Facultad en A4 horizontal
    col_widths = [5.0*cm, 70, 60, 3.6*cm, 3.6*cm, 2.7*cm, 1.8*cm, 1.8*cm, 1.8*cm, 1.8*cm, 1.2*cm]
    data = [headers_p]
    for s, a, t, esc, fac in rows:
        inicio_naive = s.hora_entrada.replace(tzinfo=None) if s.hora_entrada else None
        fin_naive    = s.hora_salida.replace(tzinfo=None)  if s.hora_salida  else None
        duracion     = str(int((fin_naive - inicio_naive).total_seconds() / 60)) if (inicio_naive and fin_naive) else "—"
        data.append([
            Paragraph(a.nombre, cell_st),
            Paragraph(a.codigo or "", cell_st),
            Paragraph(a.dni, cell_st),
            Paragraph(fac.nombre if fac else "", fac_st),
            Paragraph(esc.nombre if esc else "", fac_st),
            Paragraph(s.razon_uso or "—", cell_st),
            Paragraph(t.nombre_red, cell_st),
            inicio_naive.strftime("%I:%M %p") if inicio_naive else "",
            fin_naive.strftime("%I:%M %p")    if fin_naive    else "En curso",
            inicio_naive.strftime("%d/%m/%Y") if inicio_naive else "",
            duracion,
        ])

    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0),  c_azul),
        ("TEXTCOLOR",     (0, 0), (-1, 0),  colors.white),
        ("FONTNAME",      (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, 0),  7),
        ("ALIGN",         (0, 0), (-1, 0),  "CENTER"),
        ("BOTTOMPADDING", (0, 0), (-1, 0),  5),
        ("TOPPADDING",    (0, 0), (-1, 0),  5),
        ("FONTNAME",      (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE",      (0, 1), (-1, -1), 6.5),
        ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.white, c_claro]),
        ("GRID",          (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",    (0, 1), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 4),
        ("LEFTPADDING",   (0, 0), (-1, -1), 4),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 4),
    ]))
    elementos.append(table)

    import asyncio
    await asyncio.to_thread(doc.build, elementos)
    buf.seek(0)
    filename = f"historial_sesiones_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return StreamingResponse(
        buf, media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ── Dashboard Stats ─────────────────────────────────────────────────

@router.get("/dashboard/stats")
async def dashboard_stats(
    db: AsyncSession = Depends(get_db),
    _: Usuario = Depends(obtener_usuario_actual)
):
    """Estadísticas para el panel de administración."""
    from sqlalchemy import func

    # Total terminales y estados
    terminales = await db.execute(select(func.count(Terminal.id)))
    total_terminales = terminales.scalar()

    activas_q = await db.execute(select(func.count(Terminal.id)).where(Terminal.estado == "activo"))
    terminales_activas = activas_q.scalar()

    # Sesiones activas
    sesiones_q = await db.execute(select(func.count(Sesion.id)).where(Sesion.estado == "activa"))
    sesiones_activas_count = sesiones_q.scalar()

    # Total alumnos
    alumnos_q = await db.execute(select(func.count(AlumnoMaestro.dni)))
    total_alumnos = alumnos_q.scalar()

    return {
        "total_terminales": total_terminales,
        "terminales_activas": terminales_activas,
        "sesiones_activas": sesiones_activas_count,
        "total_alumnos": total_alumnos
    }


@router.post("/admin/cerrar-todas")
async def cerrar_todas_sesiones(
    db: AsyncSession = Depends(get_db),
    admin: Usuario = Depends(obtener_usuario_actual)
):
    """Cierra todas las sesiones activas en el sistema."""
    if admin.rol != "superadmin":
        raise HTTPException(status_code=403, detail="No tiene permisos para esta acción")
    
    res = await db.execute(select(Sesion).where(Sesion.estado == "activa"))
    sesiones = res.scalars().all()

    ahora = datetime.now().replace(tzinfo=None)   # BUG-1: era utcnow() → desfase de 5h en Perú
    for s in sesiones:
        s.activa        = False
        s.hora_salida   = ahora
        s.motivo_cierre = "admin_bulk"

    # Las PCs vuelven a la pantalla de login → estado "bloqueado" (G-12),
    # coherente con el resto del sistema (antes usaba "disponible", estado que
    # ninguna otra parte produce).
    await db.execute(update(Terminal).values(estado="bloqueado"))

    await registrar_auditoria(admin.username, "cerrar_todas_sesiones", rol=admin.rol,
                              detalle=f"{len(sesiones)} sesión(es) cerrada(s)", db=db)
    await db.commit()

    # Indicar a los clientes WPF que vuelvan a la pantalla de login
    await manager.forzar_cierre_sesion_todas()
    await manager.notificar_evento(f"Administrador finalizó TODAS las sesiones activas ({len(sesiones)} sesiones)", "warning")
    await manager.notificar_admins()

    return {"mensaje": f"Se han cerrado {len(sesiones)} sesiones correctamente"}


@router.delete("/admin/limpiar-sesiones")
async def limpiar_sesiones(
    db: AsyncSession = Depends(get_db),
    admin: Usuario = Depends(obtener_usuario_actual)
):
    """Cierra sesiones activas y borra todo el historial. Mantiene terminales y alumnos."""
    from main import logger
    logger.info(f"[ADMIN] Usuario '{admin.username}' solicitó LIMPIAR HISTORIAL DE SESIONES")

    if admin.rol != "superadmin":
        raise HTTPException(status_code=403, detail="No tiene permisos para esta acción")

    # Primero notificar a los clientes WPF que vuelvan al login antes de borrar
    await manager.notificar_evento("LIMPIAR HISTORIAL: Cerrando sesiones activas...", "warning")
    await manager.forzar_cierre_sesion_todas()

    # Borrar solo el historial de sesiones; terminales y alumnos se mantienen
    await db.execute(delete(Sesion))
    # Las PCs quedan en la pantalla de login → estado "bloqueado" (G-12)
    await db.execute(update(Terminal).values(estado="bloqueado"))
    await registrar_auditoria(admin.username, "limpiar_historial_sesiones", rol=admin.rol,
                              detalle="borró todo el historial de sesiones", db=db)
    await db.commit()

    await manager.notificar_evento("Historial borrado. Terminales listas para nuevo periodo.", "warning")
    return {"mensaje": "Historial de sesiones borrado correctamente. Terminales y alumnos mantenidos."}


@router.post("/admin/importar-excel")
async def importar_excel_upload(
    archivo: UploadFile,
    db: AsyncSession = Depends(get_db),
    admin: Usuario = Depends(obtener_usuario_actual)
):
    """Importa historial de sesiones desde un Excel con el formato de exportación estándar."""
    from main import logger
    import io
    import openpyxl

    if admin.rol != "superadmin":
        raise HTTPException(status_code=403, detail="No tiene permisos para esta acción")

    logger.info(f"[ADMIN] '{admin.username}' inició importación de Excel")

    contenido = await _leer_upload_limitado(archivo)  # E-10: límite de tamaño
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        from main import logger
        logger.warning(f"[IMPORT-EXCEL] Archivo inválido: {e}")
        raise HTTPException(status_code=400, detail="El archivo no es un Excel válido (.xlsx).")

    # ── Obtener o crear terminal virtual para sesiones importadas ──
    # Se hace commit inmediato para que no se pierda si hay rollback en filas posteriores
    try:
        res_vt = await db.execute(select(Terminal).where(Terminal.nombre_red == "IMPORTADO"))
        terminal_virtual = res_vt.scalar_one_or_none()
        if not terminal_virtual:
            terminal_virtual = Terminal(nombre_red="IMPORTADO", ip="0.0.0.0", estado="offline")
            db.add(terminal_virtual)
            await db.commit()  # commit inmediato, nunca habrá duplicate entry
            # Re-fetch para tener id correcto en sesión fresca
            res_vt2 = await db.execute(select(Terminal).where(Terminal.nombre_red == "IMPORTADO"))
            terminal_virtual = res_vt2.scalar_one_or_none()
        terminal_virtual_id = terminal_virtual.id
    except Exception as e:
        await db.rollback()
        from main import logger
        logger.error(f"[IMPORT-EXCEL] Error preparando terminal virtual: {e}")
        raise HTTPException(status_code=500, detail="Error al preparar la importación. Revise los logs.")

    # ── Mapear encabezados (case-insensitive, sin espacios) ──
    headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]

    def col(variantes):
        for v in variantes:
            if v in headers:
                return headers.index(v)
        return None

    idx_dni        = col(["dni"])
    idx_codigo     = col(["código", "codigo", "cod", "código universitario", "codigo universitario", "código de matrícula", "matricula"])
    idx_estudiante = col(["estudiante", "nombre completo", "nombre_completo", "apellidos y nombres", "apellidos nombres", "nombre", "nombres", "alumno"])
    idx_actividad  = col(["actividad", "razón", "razon", "motivo", "actividad uso"])
    idx_inicio     = col(["inicio", "hora entrada", "hora_entrada", "entrada"])
    idx_salida     = col(["salida", "hora salida", "hora_salida", "fin"])
    idx_fecha      = col(["fecha", "fecha uso", "fecha_uso", "fecha de uso"])

    if idx_dni is None:
        raise HTTPException(status_code=400, detail="No se encontró la columna 'DNI' en el Excel. Verifique el formato.")

    def cell(fila, idx):
        return str(fila[idx]).strip() if idx is not None and idx < len(fila) and fila[idx] is not None else ""

    def parse_dt(s):
        for fmt in ("%H:%M:%S", "%H:%M", "%I:%M:%S %p", "%I:%M %p"):
            try:
                return datetime.strptime(s, fmt)
            except Exception:
                pass
        return None

    def parse_fecha(s):
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(s, fmt).date()
            except Exception:
                pass
        return datetime.now().date()

    insertadas = 0
    errores    = 0
    errores_detalle = []

    for num_fila, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Saltar filas completamente vacías
        if all(v is None for v in fila):
            continue
        try:
            dni        = cell(fila, idx_dni)
            codigo     = cell(fila, idx_codigo)
            estudiante = cell(fila, idx_estudiante)
            actividad  = cell(fila, idx_actividad)
            inicio_s   = cell(fila, idx_inicio)
            salida_s   = cell(fila, idx_salida)
            fecha_s    = cell(fila, idx_fecha)

            # Validar DNI
            dni_limpio = dni.replace(" ", "").replace("-", "")
            if not dni_limpio.isdigit() or len(dni_limpio) != 8:
                errores += 1
                errores_detalle.append(f"Fila {num_fila}: DNI inválido '{dni}'")
                continue

            # Buscar alumno por DNI, luego por código
            res_a = await db.execute(select(AlumnoMaestro).where(AlumnoMaestro.dni == dni_limpio))
            alumno = res_a.scalar_one_or_none()

            if not alumno and codigo:
                res_a2 = await db.execute(select(AlumnoMaestro).where(AlumnoMaestro.codigo == codigo))
                alumno = res_a2.scalar_one_or_none()

            if not alumno:
                alumno = AlumnoMaestro(
                    dni=dni_limpio,
                    codigo=codigo or None,
                    nombre=estudiante or "SIN NOMBRE",
                )
                db.add(alumno)
                await db.flush()

            fecha_obj  = parse_fecha(fecha_s)
            inicio_obj = parse_dt(inicio_s)
            salida_obj = parse_dt(salida_s)

            inicio_dt = datetime.combine(fecha_obj, inicio_obj.time()) if inicio_obj else datetime.combine(fecha_obj, datetime.min.time())
            salida_dt = datetime.combine(fecha_obj, salida_obj.time()) if salida_obj else None

            sesion = Sesion(
                dni_alumno=alumno.dni,
                id_terminal=terminal_virtual_id,
                hora_entrada=inicio_dt,
                hora_salida=salida_dt,
                estado="cerrada",
                motivo_cierre="importacion_excel",
                razon_uso=actividad or None,
                fecha_uso=fecha_obj,
            )
            db.add(sesion)
            insertadas += 1

        except Exception as ex:
            # Rollback solo la fila problemática; el terminal_virtual ya está commiteado
            try:
                await db.rollback()
            except Exception:
                pass
            logger.warning(f"[IMPORT] Error en fila {num_fila}: {ex}")
            errores += 1
            errores_detalle.append(f"Fila {num_fila}: {ex}")
            # terminal_virtual_id ya existe en DB (commiteado antes), solo re-verificamos
            try:
                res_vt2 = await db.execute(select(Terminal).where(Terminal.nombre_red == "IMPORTADO"))
                tv2 = res_vt2.scalar_one_or_none()
                if tv2:
                    terminal_virtual_id = tv2.id
            except Exception:
                pass

    try:
        await db.commit()
    except Exception as e:
        await db.rollback()
        from main import logger
        logger.error(f"[IMPORT-EXCEL] Error al guardar en BD: {e}")
        raise HTTPException(status_code=500, detail="Error al guardar en la base de datos. Revise los logs.")

    msg = f"Importación completada: {insertadas} registro(s) insertado(s)"
    if errores:
        msg += f", {errores} fila(s) ignorada(s)"
    logger.info(f"[ADMIN] {msg}")
    await manager.notificar_evento(f"{msg}", "warning")
    return {
        "mensaje": msg,
        "insertadas": insertadas,
        "errores": errores,
        "detalle_errores": errores_detalle[:10]  # máx 10 para no saturar la respuesta
    }


@router.post("/admin/importar-maestro")
async def importar_maestro(
    archivo: UploadFile,
    db: AsyncSession = Depends(get_db),
    admin: Usuario = Depends(obtener_usuario_actual)
):
    """Importación incremental del maestro de alumnos desde Excel. Upsert por DNI."""
    from main import logger
    import io
    import openpyxl

    if admin.rol != "superadmin":
        raise HTTPException(status_code=403, detail="No tiene permisos para esta acción")

    contenido = await _leer_upload_limitado(archivo)  # E-10: límite de tamaño
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        from main import logger
        logger.warning(f"[IMPORT-EXCEL] Archivo inválido: {e}")
        raise HTTPException(status_code=400, detail="El archivo no es un Excel válido (.xlsx).")

    import unicodedata

    def _norm(s: str) -> str:
        """Minúsculas + sin tildes + sin puntos para comparación flexible."""
        s = s.lower().replace(".", "").strip()
        return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")

    raw_headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    headers_norm = [_norm(h) for h in raw_headers]

    def col(variantes):
        for v in variantes:
            vn = _norm(v)
            for i, h in enumerate(headers_norm):
                if vn == h or h.startswith(vn):
                    return i
        return None

    def col_nombre_real(variantes) -> str:
        """Devuelve el nombre real de la columna encontrada (para logs)."""
        for v in variantes:
            vn = _norm(v)
            for i, h in enumerate(headers_norm):
                if vn == h or h.startswith(vn):
                    return raw_headers[i]
        return ""

    idx_dni       = col(["dni"])
    idx_apellidos = col(["apellidos", "a_paterno a_materno", "apellidos completos"])
    idx_paterno   = col(["a_paterno", "apellido paterno", "paterno", "primer apellido"])
    idx_materno   = col(["a_materno", "apellido materno", "materno", "segundo apellido"])
    idx_nombres_p = col(["nombres"])
    idx_nombre    = col(["nombre_completo", "nombre completo", "apellidos y nombres", "apellidos nombres", "apellidos y nombres del estudiante", "estudiante", "nombre", "alumno"])
    idx_codigo    = col(["codigo_universitario", "codigo universitario", "codigo", "cod", "codigo de matricula", "matricula"])
    idx_facultad  = col(["facultad", "nombre facultad", "fac"])
    idx_escuela   = col(["escuela", "escuela profesional", "carrera"])

    # Log de columnas detectadas
    logger.info(f"[IMPORT-MAESTRO] DNI col={raw_headers[idx_dni] if idx_dni is not None else 'NO'} | "
                f"FAC col={col_nombre_real(['facultad','nombre facultad','fac']) or 'NO'} | "
                f"ESC col={col_nombre_real(['escuela','escuela profesional','carrera']) or 'NO'}")

    # Modo nombre: APELLIDOS+NOMBRES | A_PATERNO+A_MATERNO+NOMBRES | columna única
    if idx_apellidos is not None and idx_nombres_p is not None:
        _modo_nombre = "apellidos_nombres"
    elif idx_paterno is not None and idx_nombres_p is not None:
        _modo_nombre = "partes"
    else:
        _modo_nombre = "unico"

    if idx_dni is None:
        raise HTTPException(status_code=400, detail="Columna 'dni' no encontrada en el Excel.")

    def cell(fila, idx):
        return str(fila[idx]).strip() if idx is not None and idx < len(fila) and fila[idx] is not None else ""

    def titulo(s: str) -> str:
        return s.title().strip() if s else ""

    def limpiar_escuela(escuela: str) -> str | None:
        if not escuela:
            return None
        e = escuela.replace(".", "").strip().upper()
        if not e:
            return None
        return e

    def limpiar_facultad(s: str) -> str:
        return s.replace(".", "").strip().upper() if s else ""

    # Cache en memoria para evitar consultas repetidas por facultad/escuela
    _cache_fac: dict[str, int] = {}
    _cache_esc: dict[tuple, int] = {}

    async def _get_or_create_facultad(nombre_fac: str) -> int | None:
        if not nombre_fac:
            return None
        if nombre_fac in _cache_fac:
            return _cache_fac[nombre_fac]
        r = await db.execute(select(Facultad).where(Facultad.nombre == nombre_fac))
        fac = r.scalar_one_or_none()
        if not fac:
            fac = Facultad(nombre=nombre_fac)
            db.add(fac)
            await db.flush()
        _cache_fac[nombre_fac] = fac.id
        return fac.id

    async def _get_or_create_escuela(nombre_esc: str, id_fac: int | None) -> int | None:
        if not nombre_esc:
            return None
        key = (nombre_esc, id_fac)
        if key in _cache_esc:
            return _cache_esc[key]
        q_esc = select(Escuela).where(Escuela.nombre == nombre_esc)
        if id_fac:
            q_esc = q_esc.where(Escuela.id_facultad == id_fac)
        r = await db.execute(q_esc)
        esc = r.scalar_one_or_none()
        if not esc:
            esc = Escuela(nombre=nombre_esc, id_facultad=id_fac)
            db.add(esc)
            await db.flush()
        _cache_esc[key] = esc.id
        return esc.id

    insertados = actualizados = errores = 0
    errores_detalle = []

    for num_fila, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None for v in fila):
            continue
        try:
            dni = cell(fila, idx_dni).replace(" ", "").replace("-", "")
            if not dni.isdigit() or len(dni) != 8:
                errores += 1
                errores_detalle.append(f"Fila {num_fila}: DNI inválido '{dni}'")
                continue

            if _modo_nombre == "apellidos_nombres":
                apellidos = titulo(cell(fila, idx_apellidos))
                nombres_v = titulo(cell(fila, idx_nombres_p))
                nombre = f"{apellidos} {nombres_v}".strip()
            elif _modo_nombre == "partes":
                paterno = titulo(cell(fila, idx_paterno))
                materno = titulo(cell(fila, idx_materno) if idx_materno is not None else "")
                nombres_v = titulo(cell(fila, idx_nombres_p))
                apellidos = " ".join(p for p in [paterno, materno] if p)
                nombre = f"{apellidos} {nombres_v}".strip()
            else:
                nombre = titulo(cell(fila, idx_nombre))
            codigo        = cell(fila, idx_codigo) or None
            nombre_fac    = limpiar_facultad(cell(fila, idx_facultad))
            nombre_esc_raw = cell(fila, idx_escuela).strip()
            nombre_esc     = limpiar_escuela(nombre_esc_raw) or ""

            # Fase 1: Facultad
            id_fac = await _get_or_create_facultad(nombre_fac)
            # Fase 2: Escuela — SOLO si pasó el filtro de limpieza
            id_esc = await _get_or_create_escuela(nombre_esc, id_fac) if nombre_esc else None

            # Fase 3: Upsert AlumnoMaestro
            res = await db.execute(select(AlumnoMaestro).where(AlumnoMaestro.dni == dni))
            existente = res.scalar_one_or_none()

            if existente:
                if nombre:              existente.nombre      = nombre
                if codigo:              existente.codigo      = codigo
                if id_esc is not None:  existente.id_escuela  = id_esc
                if id_fac is not None:  existente.id_facultad = id_fac
                actualizados += 1
            else:
                db.add(AlumnoMaestro(
                    dni=dni,
                    nombre=nombre or "SIN NOMBRE",
                    codigo=codigo or None,
                    id_escuela=id_esc,
                    id_facultad=id_fac,
                ))
                insertados += 1

            # Flush cada 200 filas para liberar memoria sin hacer commit parcial
            if (insertados + actualizados) % 200 == 0:
                await db.flush()

        except Exception as ex:
            errores += 1
            errores_detalle.append(f"Fila {num_fila}: {ex}")

    await registrar_auditoria(admin.username, "importar_maestro", rol=admin.rol,
                              objetivo="padrón de alumnos",
                              detalle=f"{insertados} nuevo(s), {actualizados} actualizado(s), {errores} ignorada(s)", db=db)
    await db.commit()
    msg = f"Maestro importado: {insertados} alumno(s) nuevo(s), {actualizados} actualizado(s)"
    if errores:
        msg += f", {errores} fila(s) ignorada(s)"
    logger.info(f"[ADMIN] {msg}")
    return {"mensaje": msg, "insertados": insertados, "actualizados": actualizados,
            "errores": errores, "detalle_errores": errores_detalle[:10]}


@router.get("/admin/maestro")
async def listar_maestro(
    db: AsyncSession = Depends(get_db),
    admin: Usuario = Depends(obtener_usuario_actual),
    search: str | None = None,
    limit: int = 50,
    offset: int = 0,
):
    """Lista paginada del maestro de alumnos con búsqueda opcional."""
    if admin.rol not in ("superadmin", "admin"):
        raise HTTPException(status_code=403, detail="Acceso denegado")
    from sqlalchemy import or_, func
    from sqlalchemy.orm import aliased
    FacDir2 = aliased(Facultad, name="fac_direct2")

    q = (select(AlumnoMaestro, Escuela, Facultad, FacDir2)
         .outerjoin(Escuela,   AlumnoMaestro.id_escuela  == Escuela.id)
         .outerjoin(Facultad,  Escuela.id_facultad       == Facultad.id)
         .outerjoin(FacDir2,   AlumnoMaestro.id_facultad == FacDir2.id))
    if search:
        like = f"%{search}%"
        q = q.where(or_(
            AlumnoMaestro.dni.ilike(like),
            AlumnoMaestro.nombre.ilike(like),
            AlumnoMaestro.codigo.ilike(like),
        ))
    total_q = select(func.count()).select_from(q.subquery())
    total = (await db.execute(total_q)).scalar()
    limit = min(max(limit, 1), 200)   # G-13: tope de seguridad
    q = q.order_by(AlumnoMaestro.nombre).offset(offset).limit(limit)
    rows = (await db.execute(q)).all()
    return {
        "total": total,
        "alumnos": [
            {
                "dni":        r.dni,
                "nombre":     r.nombre,
                "codigo":     r.codigo,
                "facultad":   (fac.nombre if fac else None) or (fd.nombre if fd else ""),
                "escuela":    esc.nombre if esc else "",
                "id_escuela": r.id_escuela,
            }
            for r, esc, fac, fd in rows
        ]
    }


class AlumnoMaestroUpdate(BaseModel):
    nombre:     str | None = None
    codigo:     str | None = None
    facultad:   str | None = None
    escuela:    str | None = None


class AlumnoMaestroNuevo(BaseModel):
    dni:      str
    nombre:   str
    codigo:   str | None = None
    facultad: str | None = None
    escuela:  str | None = None


@router.post("/admin/maestro/nuevo")
async def crear_usuario_manual(
    datos: AlumnoMaestroNuevo,
    db: AsyncSession = Depends(get_db),
    admin: Usuario = Depends(obtener_usuario_actual),
):
    """Registra manualmente un nuevo usuario en el maestro."""
    if admin.rol not in ("superadmin", "admin"):
        raise HTTPException(status_code=403, detail="Acceso denegado")

    if not datos.dni.isdigit() or len(datos.dni) != 8:
        raise HTTPException(status_code=422, detail="El DNI debe tener exactamente 8 dígitos numéricos")
    if not datos.nombre.strip():
        raise HTTPException(status_code=422, detail="El nombre completo es obligatorio")

    res = await db.execute(select(AlumnoMaestro).where(AlumnoMaestro.dni == datos.dni))
    if res.scalar_one_or_none():
        raise HTTPException(status_code=409, detail=f"Ya existe un usuario con DNI {datos.dni}")

    # Resolver o crear facultad/escuela
    id_fac = None
    id_esc = None
    if datos.facultad:
        nombre_fac = datos.facultad.replace(".", "").strip().upper()
        r = await db.execute(select(Facultad).where(Facultad.nombre == nombre_fac))
        fac = r.scalar_one_or_none()
        if not fac:
            fac = Facultad(nombre=nombre_fac)
            db.add(fac)
            await db.flush()
        id_fac = fac.id

    if datos.escuela:
        nombre_esc = datos.escuela.replace(".", "").strip().upper()
        q_esc = select(Escuela).where(Escuela.nombre == nombre_esc)
        if id_fac:
            q_esc = q_esc.where(Escuela.id_facultad == id_fac)
        r2 = await db.execute(q_esc)
        esc = r2.scalar_one_or_none()
        if not esc:
            esc = Escuela(nombre=nombre_esc, id_facultad=id_fac)
            db.add(esc)
            await db.flush()
        id_esc = esc.id

    alumno = AlumnoMaestro(
        dni=datos.dni,
        nombre=datos.nombre.strip(),
        codigo=datos.codigo.strip() if datos.codigo else None,
        id_facultad=id_fac,
        id_escuela=id_esc,
    )
    db.add(alumno)
    await registrar_auditoria(admin.username, "crear_alumno", rol=admin.rol,
                              objetivo=f"DNI {datos.dni}", detalle=datos.nombre.strip(), db=db)
    await db.commit()
    return {"mensaje": f"Usuario '{datos.nombre.strip()}' registrado correctamente"}


@router.put("/admin/maestro/{dni}")
async def actualizar_maestro(
    dni: str,
    datos: AlumnoMaestroUpdate,
    db: AsyncSession = Depends(get_db),
    admin: Usuario = Depends(obtener_usuario_actual),
):
    """Edición manual de un registro del maestro por DNI."""
    if admin.rol != "superadmin":
        raise HTTPException(status_code=403, detail="Solo administradores pueden editar el maestro")
    res = await db.execute(select(AlumnoMaestro).where(AlumnoMaestro.dni == dni))
    alumno = res.scalar_one_or_none()
    if not alumno:
        raise HTTPException(status_code=404, detail="Alumno no encontrado en el maestro")
    if datos.nombre is not None:
        alumno.nombre = datos.nombre.strip()
    if datos.codigo is not None:
        alumno.codigo = datos.codigo.strip() or None
    if datos.facultad is not None or datos.escuela is not None:
        fac_nombre = datos.facultad.replace(".", "").strip().upper() if datos.facultad else None
        esc_nombre = datos.escuela.replace(".", "").strip().upper() if datos.escuela else None
        id_fac = None
        if fac_nombre:
            r = await db.execute(select(Facultad).where(Facultad.nombre == fac_nombre))
            fac_obj = r.scalar_one_or_none()
            if not fac_obj:
                fac_obj = Facultad(nombre=fac_nombre)
                db.add(fac_obj)
                await db.flush()
            id_fac = fac_obj.id
        id_esc = None
        if esc_nombre and id_fac:
            r = await db.execute(select(Escuela).where(Escuela.nombre == esc_nombre, Escuela.id_facultad == id_fac))
            esc_obj = r.scalar_one_or_none()
            if not esc_obj:
                esc_obj = Escuela(nombre=esc_nombre, id_facultad=id_fac)
                db.add(esc_obj)
                await db.flush()
            id_esc = esc_obj.id
        # BUG-2: asignar id_facultad (antes nunca se actualizaba) y solo tocar
        # id_escuela cuando realmente se resolvió una escuela — así editar solo
        # la facultad no borra la escuela existente del alumno.
        if id_fac is not None:
            alumno.id_facultad = id_fac
        if id_esc is not None:
            alumno.id_escuela = id_esc
    await registrar_auditoria(admin.username, "editar_alumno", rol=admin.rol,
                              objetivo=f"DNI {dni}", db=db)
    await db.commit()
    return {"mensaje": f"Alumno {dni} actualizado correctamente"}


@router.delete("/admin/maestro/{dni}")
async def eliminar_maestro(
    dni: str,
    db: AsyncSession = Depends(get_db),
    admin: Usuario = Depends(obtener_usuario_actual),
):
    """Elimina un registro del maestro por DNI, incluyendo todas sus sesiones."""
    if admin.rol != "superadmin":
        raise HTTPException(status_code=403, detail="Solo administradores pueden eliminar del maestro")
    res = await db.execute(select(AlumnoMaestro).where(AlumnoMaestro.dni == dni))
    alumno = res.scalar_one_or_none()
    if not alumno:
        raise HTTPException(status_code=404, detail="Alumno no encontrado")
    
    # Eliminar todas las sesiones del alumno primero para evitar violación de llave foránea
    await db.execute(delete(Sesion).where(Sesion.dni_alumno == dni))

    # Luego eliminar el alumno
    await db.delete(alumno)
    await registrar_auditoria(admin.username, "eliminar_alumno", rol=admin.rol,
                              objetivo=f"DNI {dni}",
                              detalle="eliminado del maestro con todas sus sesiones", db=db)
    await db.commit()
    return {"mensaje": f"Alumno {dni} eliminado del maestro con todas sus sesiones"}


@router.delete("/admin/reset-maestro")
async def reset_maestro(
    db: AsyncSession = Depends(get_db),
    admin: Usuario = Depends(obtener_usuario_actual)
):
    """Elimina todos los alumnos del maestro y sus sesiones. Mantiene terminales y usuarios."""
    from main import logger
    logger.info(f"[ADMIN] Usuario '{admin.username}' solicitó LIMPIAR MAESTRO DE ALUMNOS")

    if admin.rol != "superadmin":
        raise HTTPException(status_code=403, detail="No tiene permisos para esta acción")

    res = await db.execute(select(Sesion))
    total_sesiones = len(res.scalars().all())
    res2 = await db.execute(select(AlumnoMaestro))
    total_alumnos = len(res2.scalars().all())

    # Eliminar sesiones primero (FK), luego alumnos
    await db.execute(delete(Sesion))
    await db.execute(delete(AlumnoMaestro))
    await db.execute(update(Terminal).values(estado="bloqueado"))
    await registrar_auditoria(admin.username, "reset_maestro", rol=admin.rol,
                              detalle=f"borró {total_alumnos} alumno(s) y {total_sesiones} sesión(es)", db=db)
    await db.commit()

    await manager.forzar_cierre_sesion_todas()
    await manager.notificar_evento(
        f"BASE DE DATOS LIMPIADA: {total_alumnos} alumno(s) y {total_sesiones} sesión(es) eliminados por '{admin.username}'",
        "warning"
    )
    await manager.notificar_admins()

    return {"mensaje": f"Base de datos limpiada: {total_alumnos} alumno(s) y {total_sesiones} sesión(es) eliminados. Terminales y usuarios administradores conservados."}


class BanCreate(BaseModel):
    dni: str
    motivo: str
    dias: int = 7


class PersonalCreate(BaseModel):
    dni:      str
    nombre:   str
    cargo:    str | None = None
    area:     str | None = None
    correo:   str | None = None
    telefono: str | None = None


class PersonalUpdate(BaseModel):
    nombre:   str | None = None
    cargo:    str | None = None
    area:     str | None = None
    correo:   str | None = None
    telefono: str | None = None


@router.post("/admin/bans")
async def crear_ban(
    datos: BanCreate,
    db: AsyncSession = Depends(get_db),
    admin: Usuario = Depends(obtener_usuario_actual),
):
    """Banea a un alumno por N días (solo admin)."""
    if admin.rol != "superadmin":
        raise HTTPException(status_code=403, detail="Solo administradores pueden banear usuarios")
    res = await db.execute(select(AlumnoMaestro).where(AlumnoMaestro.dni == datos.dni))
    if not res.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="Alumno no encontrado")
    from datetime import timedelta
    ahora = datetime.now()
    fecha_fin = ahora + timedelta(days=datos.dias)
    if not datos.motivo or not datos.motivo.strip():
        raise HTTPException(status_code=422, detail="El motivo del ban es obligatorio")
    ban = Ban(
        dni_alumno=datos.dni,
        motivo=datos.motivo.strip(),
        fecha_ini=ahora,
        fecha_fin=fecha_fin,
        baneado_por=admin.username,
    )
    db.add(ban)
    await registrar_auditoria(admin.username, "banear_alumno", rol=admin.rol,
                              objetivo=f"DNI {datos.dni}",
                              detalle=f"{datos.dias} día(s) — {datos.motivo.strip()}", db=db)
    await db.commit()
    return {"mensaje": f"Alumno {datos.dni} baneado hasta {fecha_fin.strftime('%d/%m/%Y')}", "fecha_fin": fecha_fin}


@router.delete("/admin/bans/{ban_id}")
async def eliminar_ban(
    ban_id: int,
    db: AsyncSession = Depends(get_db),
    admin: Usuario = Depends(obtener_usuario_actual),
):
    """Levanta un ban por su ID y resetea las incidencias activas del alumno (solo admin)."""
    if admin.rol != "superadmin":
        raise HTTPException(status_code=403, detail="Solo administradores pueden levantar bans")
    res = await db.execute(select(Ban).where(Ban.id == ban_id))
    ban = res.scalar_one_or_none()
    if not ban:
        raise HTTPException(status_code=404, detail="Ban no encontrado")
    ahora = datetime.now()
    ban.levantado_por  = admin.username
    ban.fecha_levantado = ahora
    ban.fecha_fin      = ahora  # expirar inmediatamente
    # Resetear todas las incidencias activas del alumno
    await db.execute(
        update(Incidencia)
        .where(Incidencia.dni_alumno == ban.dni_alumno, Incidencia.activa == True)
        .values(activa=False)
    )
    await registrar_auditoria(admin.username, "levantar_ban", rol=admin.rol,
                              objetivo=f"DNI {ban.dni_alumno}", db=db)
    await db.commit()
    return {"mensaje": "Ban levantado y incidencias reseteadas correctamente"}


@router.get("/admin/bans")
async def listar_bans(
    db: AsyncSession = Depends(get_db),
    admin: Usuario = Depends(obtener_usuario_actual),
):
    """Lista todos los bans activos (no expirados)."""
    if admin.rol != "superadmin":
        raise HTTPException(status_code=403, detail="Solo administradores pueden ver bans")
    from sqlalchemy.orm import joinedload
    ahora = datetime.now()
    res = await db.execute(
        select(Ban)
        .options(joinedload(Ban.alumno))
        .where((Ban.fecha_fin == None) | (Ban.fecha_fin > ahora))
        .order_by(Ban.fecha_ini.desc())
    )
    bans = res.scalars().all()
    return [
        {
            "id":          b.id,
            "dni":         b.dni_alumno,
            "nombre":      b.alumno.nombre if b.alumno else b.dni_alumno,
            "motivo":      b.motivo or "",
            "fecha_ini":   b.fecha_ini,
            "fecha_fin":   b.fecha_fin,
            "baneado_por": b.baneado_por or "",
        }
        for b in bans
    ]


@router.get("/admin/bans/historial")
async def historial_bans(
    db: AsyncSession = Depends(get_db),
    admin: Usuario = Depends(obtener_usuario_actual),
    dni: str | None = None,
    estado: str | None = None,   # activo | expirado | levantado
    limit: int = 50,
    offset: int = 0,
):
    """Historial completo de bans (activos, expirados y levantados manualmente)."""
    if admin.rol != "superadmin":
        raise HTTPException(status_code=403, detail="Solo el superadmin puede ver el historial de bans")
    from sqlalchemy.orm import joinedload
    from sqlalchemy import func as _func
    limit = min(max(limit, 1), 200)   # G-13: tope de seguridad
    ahora = datetime.now()
    q = select(Ban).options(joinedload(Ban.alumno)).order_by(Ban.fecha_ini.desc())
    if dni:
        q = q.where(Ban.dni_alumno == dni)
    if estado == "activo":
        q = q.where((Ban.fecha_fin == None) | (Ban.fecha_fin > ahora))
        q = q.where(Ban.levantado_por == None)
    elif estado == "expirado":
        q = q.where(Ban.fecha_fin <= ahora)
        q = q.where(Ban.levantado_por == None)
    elif estado == "levantado":
        q = q.where(Ban.levantado_por != None)

    total_res = await db.execute(select(_func.count()).select_from(q.subquery()))
    total = total_res.scalar_one()
    res = await db.execute(q.limit(limit).offset(offset))
    bans = res.scalars().all()

    def _estado(b: Ban) -> str:
        if b.levantado_por:
            return "levantado"
        if b.fecha_fin and b.fecha_fin <= ahora:
            return "expirado"
        return "activo"

    return {
        "total": total,
        "items": [
            {
                "id":               b.id,
                "dni":              b.dni_alumno,
                "nombre":           b.alumno.nombre if b.alumno else b.dni_alumno,
                "motivo":           b.motivo or "",
                "fecha_ini":        b.fecha_ini,
                "fecha_fin":        b.fecha_fin,
                "baneado_por":      b.baneado_por or "",
                "estado":           _estado(b),
                "levantado_por":    b.levantado_por or "",
                "fecha_levantado":  b.fecha_levantado,
            }
            for b in bans
        ],
    }


@router.get("/admin/auditoria")
async def listar_auditoria(
    db: AsyncSession = Depends(get_db),
    admin: Usuario = Depends(obtener_usuario_actual),
    usuario: str | None = None,
    accion: str | None = None,
    desde: str | None = None,    # ISO date "2026-06-01"
    hasta: str | None = None,
    limit: int = 50,
    offset: int = 0,
):
    """Bitácora de acciones administrativas (solo superadmin). Filtros: usuario,
    accion, rango de fechas; con paginación."""
    if admin.rol != "superadmin":
        raise HTTPException(status_code=403, detail="Solo el superadmin puede ver la auditoría")
    from sqlalchemy import func as _func
    limit = min(max(limit, 1), 200)   # tope de seguridad (G-13)
    q = select(Auditoria).order_by(Auditoria.fecha_hora.desc())
    if usuario:
        q = q.where(Auditoria.usuario == usuario)
    if accion:
        q = q.where(Auditoria.accion == accion)
    if desde:
        try:
            q = q.where(Auditoria.fecha_hora >= datetime.fromisoformat(desde))
        except ValueError:
            pass
    if hasta:
        try:
            q = q.where(Auditoria.fecha_hora < datetime.fromisoformat(hasta) + timedelta(days=1))
        except ValueError:
            pass
    total_res = await db.execute(select(_func.count()).select_from(q.subquery()))
    total = total_res.scalar_one()
    res = await db.execute(q.limit(limit).offset(offset))
    items = res.scalars().all()
    return {
        "total": total,
        "items": [
            {
                "id":         a.id,
                "fecha_hora": a.fecha_hora.isoformat() if a.fecha_hora else None,
                "usuario":    a.usuario,
                "rol":        a.rol or "",
                "accion":     a.accion,
                "objetivo":   a.objetivo or "",
                "detalle":    a.detalle or "",
                "ip_origen":  a.ip_origen or "",
            }
            for a in items
        ],
    }


@router.get("/admin/bans/check/{dni}")
async def check_ban(
    request: Request,
    dni: str,
    db: AsyncSession = Depends(get_db),
):
    """Verifica si un DNI tiene ban activo (usado por el WebSocket del kiosco)."""
    # E-6: mismo límite por IP que /alumnos/validar (evita enumerar quién está baneado).
    rate_limit.exigir(request, "kiosco", limite=30, ventana_seg=60)
    ahora = datetime.now()
    res = await db.execute(
        select(Ban)
        .where(Ban.dni_alumno == dni)
        .where((Ban.fecha_fin == None) | (Ban.fecha_fin > ahora))
    )
    ban = res.scalar_one_or_none()
    if ban:
        return {"baneado": True, "motivo": ban.motivo or "Acceso restringido", "fecha_fin": ban.fecha_fin}
    return {"baneado": False}


@router.get("/admin/personal")
async def listar_personal(
    db: AsyncSession = Depends(get_db),
    admin: Usuario = Depends(obtener_usuario_actual),
    search: str | None = None,
    limit: int = 50,
    offset: int = 0,
):
    """Lista paginada del personal universitario."""
    if admin.rol != "superadmin":
        raise HTTPException(status_code=403, detail="Solo administradores pueden acceder al personal")
    from sqlalchemy import or_, func
    q = select(PersonalUniversidad)
    if search:
        like = f"%{search}%"
        q = q.where(or_(
            PersonalUniversidad.dni.ilike(like),
            PersonalUniversidad.nombre.ilike(like),
            PersonalUniversidad.cargo.ilike(like),
            PersonalUniversidad.area.ilike(like),
        ))
    limit = min(max(limit, 1), 200)   # G-13: tope de seguridad
    total = (await db.execute(select(func.count()).select_from(q.subquery()))).scalar()
    q = q.order_by(PersonalUniversidad.nombre).offset(offset).limit(limit)
    rows = (await db.execute(q)).scalars().all()
    return {
        "total": total,
        "personal": [
            {
                "dni":      p.dni,
                "nombre":   p.nombre,
                "cargo":    p.cargo or "",
                "area":     p.area or "",
                "correo":   p.correo or "",
                "telefono": p.telefono or "",
                "activo":   p.activo,
            }
            for p in rows
        ]
    }


@router.post("/admin/personal/nuevo", status_code=201)
async def crear_personal(
    datos: PersonalCreate,
    db: AsyncSession = Depends(get_db),
    admin: Usuario = Depends(obtener_usuario_actual),
):
    """Registra un nuevo miembro del personal universitario."""
    if admin.rol != "superadmin":
        raise HTTPException(status_code=403, detail="Solo administradores pueden registrar personal")
    if not datos.dni.isdigit() or len(datos.dni) != 8:
        raise HTTPException(status_code=422, detail="El DNI debe tener exactamente 8 dígitos")
    if not datos.nombre.strip():
        raise HTTPException(status_code=422, detail="El nombre es obligatorio")
    res = await db.execute(select(PersonalUniversidad).where(PersonalUniversidad.dni == datos.dni))
    if res.scalar_one_or_none():
        raise HTTPException(status_code=409, detail=f"Ya existe personal con DNI {datos.dni}")
    db.add(PersonalUniversidad(
        dni=datos.dni,
        nombre=datos.nombre.strip(),
        cargo=datos.cargo.strip() if datos.cargo else None,
        area=datos.area.strip() if datos.area else None,
        correo=datos.correo.strip() if datos.correo else None,
        telefono=datos.telefono.strip() if datos.telefono else None,
    ))
    await registrar_auditoria(admin.username, "crear_personal", rol=admin.rol,
                              objetivo=f"DNI {datos.dni}", detalle=datos.nombre.strip(), db=db)
    await db.commit()
    return {"mensaje": f"Personal '{datos.nombre.strip()}' registrado correctamente"}


@router.put("/admin/personal/{dni}")
async def actualizar_personal(
    dni: str,
    datos: PersonalUpdate,
    db: AsyncSession = Depends(get_db),
    admin: Usuario = Depends(obtener_usuario_actual),
):
    """Edita un registro de personal por DNI."""
    if admin.rol != "superadmin":
        raise HTTPException(status_code=403, detail="Solo administradores pueden editar personal")
    res = await db.execute(select(PersonalUniversidad).where(PersonalUniversidad.dni == dni))
    p = res.scalar_one_or_none()
    if not p:
        raise HTTPException(status_code=404, detail="Personal no encontrado")
    if datos.nombre is not None:   p.nombre   = datos.nombre.strip()
    if datos.cargo is not None:    p.cargo    = datos.cargo.strip() or None
    if datos.area is not None:     p.area     = datos.area.strip() or None
    if datos.correo is not None:   p.correo   = datos.correo.strip() or None
    if datos.telefono is not None: p.telefono = datos.telefono.strip() or None
    await registrar_auditoria(admin.username, "editar_personal", rol=admin.rol,
                              objetivo=f"DNI {dni}", db=db)
    await db.commit()
    return {"mensaje": f"Personal {dni} actualizado correctamente"}


@router.delete("/admin/personal/{dni}")
async def eliminar_personal(
    dni: str,
    db: AsyncSession = Depends(get_db),
    admin: Usuario = Depends(obtener_usuario_actual),
):
    """Elimina un registro de personal por DNI."""
    if admin.rol != "superadmin":
        raise HTTPException(status_code=403, detail="Solo administradores pueden eliminar personal")
    res = await db.execute(select(PersonalUniversidad).where(PersonalUniversidad.dni == dni))
    p = res.scalar_one_or_none()
    if not p:
        raise HTTPException(status_code=404, detail="Personal no encontrado")
    nombre_p = p.nombre
    await db.delete(p)
    await registrar_auditoria(admin.username, "eliminar_personal", rol=admin.rol,
                              objetivo=f"DNI {dni}", detalle=nombre_p, db=db)
    await db.commit()
    return {"mensaje": f"Personal {dni} eliminado correctamente"}


@router.post("/admin/personal/importar")
async def importar_personal(
    archivo: UploadFile,
    db: AsyncSession = Depends(get_db),
    admin: Usuario = Depends(obtener_usuario_actual),
):
    """Importación masiva de personal desde Excel. Upsert por DNI."""
    if admin.rol != "superadmin":
        raise HTTPException(status_code=403, detail="No tiene permisos para esta acción")
    import io, openpyxl
    contenido = await _leer_upload_limitado(archivo)  # E-10: límite de tamaño
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        from main import logger
        logger.warning(f"[IMPORT-PERSONAL] Archivo inválido: {e}")
        raise HTTPException(status_code=400, detail="El archivo no es un Excel válido (.xlsx).")

    hdrs = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]

    def col(*opts):
        for o in opts:
            if o in hdrs: return hdrs.index(o)
        return None

    idx_dni      = col("dni")
    idx_nombre   = col("nombre completo", "nombre", "apellidos y nombres", "apellidos nombres", "trabajador")
    idx_cargo    = col("cargo", "puesto")
    idx_area     = col("área", "area", "oficina", "dependencia")
    idx_correo   = col("correo", "email", "correo electrónico", "correo electronico")
    idx_telefono = col("teléfono", "telefono", "celular")

    if idx_dni is None:
        raise HTTPException(status_code=400, detail="Columna 'DNI' no encontrada")

    def cell(fila, idx):
        return str(fila[idx]).strip() if idx is not None and idx < len(fila) and fila[idx] is not None else ""

    insertados = actualizados = errores = 0
    for num_fila, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None for v in fila): continue
        dni = cell(fila, idx_dni).replace(" ", "").replace("-", "")
        if not dni.isdigit() or len(dni) != 8:
            errores += 1; continue
        nombre = cell(fila, idx_nombre)
        if not nombre: errores += 1; continue
        res = await db.execute(select(PersonalUniversidad).where(PersonalUniversidad.dni == dni))
        existente = res.scalar_one_or_none()
        if existente:
            existente.nombre   = nombre
            existente.cargo    = cell(fila, idx_cargo)    or None
            existente.area     = cell(fila, idx_area)     or None
            existente.correo   = cell(fila, idx_correo)   or None
            existente.telefono = cell(fila, idx_telefono) or None
            actualizados += 1
        else:
            db.add(PersonalUniversidad(
                dni=dni, nombre=nombre,
                cargo=cell(fila, idx_cargo) or None,
                area=cell(fila, idx_area) or None,
                correo=cell(fila, idx_correo) or None,
                telefono=cell(fila, idx_telefono) or None,
            ))
            insertados += 1
        if (insertados + actualizados) % 200 == 0:
            await db.flush()
    await registrar_auditoria(admin.username, "importar_personal", rol=admin.rol,
                              objetivo="padrón de personal",
                              detalle=f"{insertados} nuevo(s), {actualizados} actualizado(s), {errores} ignorado(s)", db=db)
    await db.commit()
    return {"mensaje": f"{insertados} nuevo(s), {actualizados} actualizado(s), {errores} ignorado(s)",
            "insertados": insertados, "actualizados": actualizados, "errores": errores}


@router.get("/admin/personal/exportar")
async def exportar_personal(
    db: AsyncSession = Depends(get_db),
    admin: Usuario = Depends(obtener_usuario_actual),
):
    """Exporta el personal a Excel."""
    if admin.rol != "superadmin":
        raise HTTPException(status_code=403, detail="Solo administradores pueden exportar personal")
    import io, openpyxl
    from fastapi.responses import StreamingResponse
    rows = (await db.execute(select(PersonalUniversidad).order_by(PersonalUniversidad.nombre))).scalars().all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Personal Universidad"
    ws.append(["DNI", "Nombre Completo", "Cargo", "Área", "Correo", "Teléfono"])
    for p in rows:
        ws.append([p.dni, p.nombre, p.cargo or "", p.area or "", p.correo or "", p.telefono or ""])
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=personal_universidad.xlsx",
                 "Access-Control-Expose-Headers": "Content-Disposition"})


# ── Incidencias ─────────────────────────────────────────────────────

class IncidenciaCreate(BaseModel):
    dni: str
    tipo: str = "leve"          # leve | grave
    motivo: str
    descripcion: str | None = None


@router.post("/admin/incidencias", status_code=201)
async def crear_incidencia(
    datos: IncidenciaCreate,
    db: AsyncSession = Depends(get_db),
    admin: Usuario = Depends(obtener_usuario_actual),
):
    """Registra una incidencia (leve o grave) para un alumno."""
    if admin.rol not in ("superadmin", "admin"):
        raise HTTPException(status_code=403, detail="Acceso denegado")
    if datos.tipo not in ("leve", "grave"):
        raise HTTPException(status_code=422, detail="El tipo debe ser 'leve' o 'grave'")
    res = await db.execute(select(AlumnoMaestro).where(AlumnoMaestro.dni == datos.dni))
    alumno = res.scalar_one_or_none()
    if not alumno:
        raise HTTPException(status_code=404, detail="Alumno no encontrado")
    inc = Incidencia(
        dni_alumno=datos.dni,
        nombre_alumno=alumno.nombre,
        tipo=datos.tipo,
        motivo=datos.motivo.strip(),
        descripcion=datos.descripcion.strip() if datos.descripcion else None,
        registrado_por=admin.username,
    )
    db.add(inc)
    await registrar_auditoria(admin.username, "crear_incidencia", rol=admin.rol,
                              objetivo=f"DNI {datos.dni}",
                              detalle=f"{datos.tipo} — {datos.motivo.strip()}", db=db)
    await db.commit()
    # Contar incidencias activas para incluir en la respuesta
    cnt_q = await db.execute(
        select(Incidencia)
        .where(Incidencia.dni_alumno == datos.dni, Incidencia.activa == True)
    )
    total_activas = len(cnt_q.scalars().all())
    return {"mensaje": f"Incidencia registrada para {alumno.nombre}", "total_activas": total_activas}


@router.get("/admin/incidencias")
async def listar_incidencias(
    db: AsyncSession = Depends(get_db),
    admin: Usuario = Depends(obtener_usuario_actual),
    dni: str | None = None,
    solo_activas: bool = False,
    limit: int = 100,
    offset: int = 0,
):
    """Lista incidencias con filtro opcional por DNI y estado."""
    if admin.rol != "superadmin":
        raise HTTPException(status_code=403, detail="Solo administradores pueden ver incidencias")
    from sqlalchemy import func
    limit = min(max(limit, 1), 200)   # G-13: tope de seguridad
    q = select(Incidencia).order_by(Incidencia.fecha.desc())
    if dni:
        q = q.where(Incidencia.dni_alumno == dni)
    if solo_activas:
        q = q.where(Incidencia.activa == True)
    total = (await db.execute(select(func.count()).select_from(q.subquery()))).scalar()
    rows = (await db.execute(q.offset(offset).limit(limit))).scalars().all()
    return {
        "total": total,
        "items": [
            {
                "id":             i.id,
                "dni":            i.dni_alumno,
                "nombre_alumno":  i.nombre_alumno,
                "tipo":           i.tipo,
                "motivo":         i.motivo,
                "descripcion":    i.descripcion or "",
                "fecha":          i.fecha,
                "registrado_por": i.registrado_por,
                "activa":         i.activa,
            }
            for i in rows
        ],
    }


@router.get("/admin/incidencias/resumen")
async def resumen_incidencias(
    db: AsyncSession = Depends(get_db),
    admin: Usuario = Depends(obtener_usuario_actual),
):
    """Devuelve conteo de incidencias activas agrupado por alumno (para badges y alertas)."""
    if admin.rol != "superadmin":
        raise HTTPException(status_code=403, detail="Solo administradores pueden ver incidencias")
    from sqlalchemy import func
    rows = (await db.execute(
        select(Incidencia.dni_alumno, Incidencia.nombre_alumno, func.count(Incidencia.id).label("total"))
        .where(Incidencia.activa == True)
        .group_by(Incidencia.dni_alumno, Incidencia.nombre_alumno)
        .order_by(func.count(Incidencia.id).desc())
    )).all()
    return [{"dni": r.dni_alumno, "nombre": r.nombre_alumno, "total": r.total} for r in rows]


@router.delete("/admin/incidencias/{incidencia_id}")
async def eliminar_incidencia(
    incidencia_id: int,
    db: AsyncSession = Depends(get_db),
    admin: Usuario = Depends(obtener_usuario_actual),
):
    """Elimina (desactiva) una incidencia individual."""
    if admin.rol != "superadmin":
        raise HTTPException(status_code=403, detail="Solo administradores pueden eliminar incidencias")
    res = await db.execute(select(Incidencia).where(Incidencia.id == incidencia_id))
    inc = res.scalar_one_or_none()
    if not inc:
        raise HTTPException(status_code=404, detail="Incidencia no encontrada")
    inc.activa = False
    await registrar_auditoria(admin.username, "eliminar_incidencia", rol=admin.rol,
                              objetivo=f"DNI {inc.dni_alumno}",
                              detalle=f"{inc.tipo} — {inc.motivo or ''}"[:120], db=db)
    await db.commit()
    return {"mensaje": "Incidencia eliminada correctamente"}


@router.delete("/admin/reset-total")
async def reset_total(
    db: AsyncSession = Depends(get_db),
    admin: Usuario = Depends(obtener_usuario_actual)
):
    """Vaciado total de las tablas de sesiones, alumnos y terminales."""
    from main import logger
    logger.info(f"[ADMIN] Usuario '{admin.username}' solicitó RESET TOTAL")

    if admin.rol != "superadmin":
        logger.warning(f"[ADMIN] Intento de reset rechazado para usuario '{admin.username}' (no es admin)")
        raise HTTPException(status_code=403, detail="No tiene permisos para esta acción")

    # El orden importa por las claves foráneas
    await db.execute(delete(Sesion))
    await db.execute(delete(AlumnoMaestro))
    await db.execute(delete(Terminal))

    # Desconectar físicamente todas las terminales para que no se re-registren solo por heartbeat
    await manager.desconectar_todo()

    await registrar_auditoria(admin.username, "reset_total", rol=admin.rol,
                              detalle="vació sesiones, alumnos y terminales", db=db)
    await db.commit()
    await manager.notificar_evento("RESET TOTAL: El sistema ha sido reseteado por el administrador", "warning")
    return {"mensaje": "Todo el sistema ha sido limpiado correctamente"}


# ── Auto-actualización del cliente ──────────────────────────────────

RUTA_DISTRIBUCION_DEFAULT = r"C:\WinSysCache"


def _bases_distribucion_permitidas() -> list:
    """Carpetas base bajo las que se permite ubicar la distribución (E-8).

    Por defecto: la carpeta default y una variable de entorno opcional
    (RUTA_DISTRIBUCION_BASE, separada por ';' si hay varias). Evita que un
    superadmin (o una sesión suya comprometida) apunte la distribución a una
    carpeta arbitraria del sistema. Si se necesita otra carpeta, se agrega
    explícitamente aquí o por env, no se acepta cualquier ruta.
    """
    import os
    import pathlib
    bases = [pathlib.Path(RUTA_DISTRIBUCION_DEFAULT)]
    extra = os.getenv("RUTA_DISTRIBUCION_BASE", "").strip()
    if extra:
        for p in extra.split(";"):
            p = p.strip()
            if p:
                bases.append(pathlib.Path(p))
    # Normalizar todas (sin exigir que existan todavía).
    return [b.resolve() for b in bases]


def _ruta_distribucion_valida(ruta: str):
    """Normaliza la ruta y verifica que cae dentro de una base permitida (E-8).

    Devuelve la ruta normalizada (Path) si es válida; lanza HTTP 422 si no.
    """
    import pathlib
    try:
        candidata = pathlib.Path(ruta).resolve()
    except Exception:
        raise HTTPException(status_code=422, detail="Ruta inválida")
    for base in _bases_distribucion_permitidas():
        try:
            # candidata == base o está dentro de base
            if candidata == base or base in candidata.parents:
                return candidata
        except Exception:
            continue
    raise HTTPException(
        status_code=422,
        detail="La ruta no está permitida. Use la carpeta de distribución autorizada."
    )


async def _obtener_ruta_distribucion(db: AsyncSession) -> str:
    from sqlalchemy import text as _text
    res = await db.execute(_text("SELECT valor FROM ajustes_sistema WHERE clave='ruta_distribucion'"))
    row = res.fetchone()
    return row[0] if row else RUTA_DISTRIBUCION_DEFAULT


@router.get("/version")
async def obtener_version(db: AsyncSession = Depends(get_db)):
    """Devuelve la versión actual del cliente disponible para descarga."""
    import pathlib
    ruta = await _obtener_ruta_distribucion(db)
    version_file = pathlib.Path(ruta) / "version.txt"
    if not version_file.exists():
        return {"version": "0.0", "sha256": ""}
    # utf-8-sig descarta el BOM que Notepad agrega al guardar como UTF-8
    version = version_file.read_text(encoding="utf-8-sig").strip()
    # E-9: incluir el SHA-256 del exe publicado si existe, para que el cliente
    # pueda verificar la integridad de la descarga.
    sha256 = ""
    sha_file = pathlib.Path(ruta) / "cliente.sha256"
    if sha_file.exists():
        sha256 = sha_file.read_text(encoding="utf-8-sig").strip()
    return {"version": version, "sha256": sha256}



@router.get("/admin/config/ruta-distribucion")
async def obtener_ruta_distribucion_config(
    db: AsyncSession = Depends(get_db),
    admin: Usuario = Depends(obtener_usuario_actual),
):
    """Devuelve la ruta de distribución actual del cliente."""
    if admin.rol != "superadmin":
        raise HTTPException(status_code=403, detail="Solo superadmin")
    ruta = await _obtener_ruta_distribucion(db)
    return {"ruta": ruta}


@router.post("/admin/config/ruta-distribucion")
async def guardar_ruta_distribucion(
    datos: dict,
    db: AsyncSession = Depends(get_db),
    admin: Usuario = Depends(obtener_usuario_actual),
):
    """Actualiza la ruta de distribución del cliente."""
    import pathlib
    from sqlalchemy import text as _text
    if admin.rol != "superadmin":
        raise HTTPException(status_code=403, detail="Solo superadmin")
    ruta = str(datos.get("ruta", "")).strip()
    if not ruta:
        raise HTTPException(status_code=422, detail="Ruta no puede estar vacía")
    # E-8: normalizar y validar contra la whitelist de bases permitidas.
    candidata = _ruta_distribucion_valida(ruta)
    # Verificar que la carpeta existe (E-12: mensaje genérico, sin revelar la ruta absoluta).
    if not candidata.exists():
        raise HTTPException(status_code=422, detail="La carpeta indicada no existe en el servidor")
    ruta_norm = str(candidata)
    await db.execute(
        _text("INSERT INTO ajustes_sistema (clave, valor) VALUES ('ruta_distribucion', :v) ON DUPLICATE KEY UPDATE valor=:v"),
        {"v": ruta_norm}
    )
    await registrar_auditoria(admin.username, "cambiar_ruta_distribucion", rol=admin.rol,
                              objetivo=ruta_norm, db=db)
    await db.commit()
    return {"mensaje": "Ruta actualizada correctamente", "ruta": ruta_norm}


